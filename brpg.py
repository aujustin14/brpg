import sys
import os
import random
import math
import openpyxl
import pyperclip
import copy
import textwrap
import time

"""
todo:
	high priority
		- refactor combat logic
		- rework inventory
		- rework status effects
		- refactor text printing
		- better animations for text
	low priority
"""


# Prints out a message and requires the user to press enter to continue. End determines where the cursor is placed (empty space is at the end of the string, \n is on a new line).
def proceduralPrint(string, end):
	print(string, end = end)
	input()


def scrollingPrint(text):
	sentences = text.split("\n")

	for sentence in sentences:
		for word in sentence:
			for character in word:
				sys.stdout.write(character)
				sys.stdout.flush()
				if (character in [",", "."]):
					time.sleep(0.10)
		if (sentence != sentences[-1]):
			sys.stdout.write("\n")
			sys.stdout.flush()
	input()


# Returns a bar based on a desired length, the alignment (0 for left align, 1 for right align), a character for the bar, a current value, and a max value.
def generalUIBar(length, alignment, characterString, currentValue, maxValue):
	if (len(characterString) > 1):
		characterString = characterString[0]

	if (alignment == 1):
		alignment = ">"
		borderStyle = "\\"
	else:
		alignment = "<"
		borderStyle = "/"

	barPercentage = min(currentValue / maxValue, 1)
	barDisplay = characterString * int(math.ceil((length * barPercentage)))

	return (borderStyle + ("{:" + alignment + str(length) + "}").format(barDisplay) + borderStyle)


# Returns a row of the battle UI based on the provided string and the alignment (0 for left align, 1 for right align).
def battleUIRow(string, alignment):
	if (len(string) > 37):
		string = string[0:34] + "..."

	if (alignment == 0):
		currentRow = "| {:<37} |".format(string)
	else:
		currentRow = " {:>37} |".format(string)

	return currentRow


# Returns a row divider for the battle UI.
def battleUIRowDivider():
	return "+–––––––––––––––––––––––––––––––––––––––+–––––––––––––––––––––––––––––––––––––––+"


# Returns a row of the level up UI based on the provided value and the mode (0 for standard text, 1 for stat names, 2 for current stats, 3 for change in stats).
def levelUpUIRow(value, mode):
	currentRow = ""
	if (mode == 0):
		currentRow += "| " + str(value)
		currentRow += "\t" * (11 - math.ceil(len(currentRow) / 8))
		currentRow += "|"
	elif (mode == 1):
		currentRow += "|\tHP\tMP\tMA\tRA\tMD\tRD\tACC\tEVA\t\t|"
	elif (mode == 2):
		currentRow += "|"
		currentRow += "\t " + "{:q>4}".format(str(value.hp))
		currentRow += "\t " + "{:q>4}".format(str(value.mp))
		currentRow += "\t " + "{:q>4}".format(str(value.ballisticAttack))
		currentRow += "\t " + "{:q>4}".format(str(value.magicAttack))
		currentRow += "\t " + "{:q>4}".format(str(value.ballisticDefense))
		currentRow += "\t " + "{:q>4}".format(str(value.magicDefense))
		currentRow += "\t " + "{:q>4}".format(str(value.accuracy))
		currentRow += "\t " + "{:q>4}".format(str(value.evade))
		currentRow += "\t\t|"

		for i in range(len(currentRow)):
			if (currentRow[i] == "q"):
				currentRow = currentRow[0:i] + " " + currentRow[i+1:len(currentRow)]
	elif (mode == 3):
		currentRow += "|"
		currentRow += "\t+" + "{:q>4}".format(str(value.hp))
		currentRow += "\t+" + "{:q>4}".format(str(value.mp))
		currentRow += "\t+" + "{:q>4}".format(str(value.ballisticAttack))
		currentRow += "\t+" + "{:q>4}".format(str(value.magicAttack))
		currentRow += "\t+" + "{:q>4}".format(str(value.ballisticDefense))
		currentRow += "\t+" + "{:q>4}".format(str(value.magicDefense))
		currentRow += "\t+" + "{:q>4}".format(str(value.accuracy))
		currentRow += "\t+" + "{:q>4}".format(str(value.evade))
		currentRow += "\t\t|"

		for i in range(len(currentRow)):
			if (currentRow[i] == "q"):
				currentRow = currentRow[0:i] + " " + currentRow[i+1:len(currentRow)]

	return currentRow


# Rounds a value into an integer and then converts the result into another specified variable type.
def superRound(value, varType):
	return varType(round(value)) if (value != None) else None


# Clears the screen.
def clearScreen():
	if (os.name == "nt"):
		os.system("cls")
	else:
		os.system("clear")


# Formats text based on alignment and length.
def formatText(text, alignment, length):
	text = str(text)

	if (alignment == 0):
		alignment = "<"
	elif (alignment == 1):
		alignment = "^"
	elif (alignment == 2):
		alignment = ">"

	# if (len(str(text)) > length):
	# 	text = text[0:(length - 3)] + "..."
	text = textwrap.shorten(text, width = length)

	return str("{:" + alignment + str(length) + "}").format(text)


# Initialize classes #
class Stats:
	def __init__(self, every=0, hp=1, mp=1, ballisticAttack=1, magicAttack=1, ballisticDefense=1, magicDefense=1, accuracy=1, evade=1):
		if (every != 0 or hp == mp == ballisticAttack == magicAttack == ballisticDefense == magicDefense == accuracy == evade != 1):
			if (hp == mp == ballisticAttack == magicAttack == ballisticDefense == magicDefense == accuracy == evade != 1):
				every = hp
			self.hasAllStats = True
			self.hp = every
			self.mp = every
			self.ballisticAttack = every
			self.magicAttack = every
			self.ballisticDefense = every
			self.magicDefense = every
			self.accuracy = every
			self.evade = every
		else:
			self.hasAllStats = False
			self.hp = hp
			self.mp = mp
			self.ballisticAttack = ballisticAttack
			self.magicAttack = magicAttack
			self.ballisticDefense = ballisticDefense
			self.magicDefense = magicDefense
			self.accuracy = accuracy
			self.evade = evade

	def __str__(self):
		return ("hp: " + formatText(self.hp, 0, 6) + "mp: " + formatText(self.mp, 0, 6) + "ballisticAttack: " + formatText(self.ballisticAttack, 0, 6) + "ballisticAttack: " + formatText(self.ballisticAttack, 0, 6) + "magicAttack: " + formatText(self.magicAttack, 0, 6) + "ballisticDefense: " + formatText(self.ballisticDefense, 0, 6) + "magicDefense: " + formatText(self.magicDefense, 0, 6) + "accuracy: " + formatText(self.accuracy, 0, 6) + "evade: " + formatText(self.evade, 0, 6))

	def __add__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP += other
			totalMP += other
			totalBallisticAttack += other
			totalMagicAttack += other
			totalBallisticDefense += other
			totalMagicDefense += other
			totalAccuracy += other
			totalEvade += other

		elif (type(other) is Stats):
			totalHP += other.hp
			totalMP += other.mp
			totalBallisticAttack += other.ballisticAttack
			totalMagicAttack += other.magicAttack
			totalBallisticDefense += other.ballisticDefense
			totalMagicDefense += other.magicDefense
			totalAccuracy += other.accuracy
			totalEvade += other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __sub__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP -= other
			totalMP -= other
			totalBallisticAttack -= other
			totalMagicAttack -= other
			totalBallisticDefense -= other
			totalMagicDefense -= other
			totalAccuracy -= other
			totalEvade -= other

		elif (type(other) is Stats):
			totalHP -= other.hp
			totalMP -= other.mp
			totalBallisticAttack -= other.ballisticAttack
			totalMagicAttack -= other.magicAttack
			totalBallisticDefense -= other.ballisticDefense
			totalMagicDefense -= other.magicDefense
			totalAccuracy -= other.accuracy
			totalEvade -= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __mul__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP *= other
			totalMP *= other
			totalBallisticAttack *= other
			totalMagicAttack *= other
			totalBallisticDefense *= other
			totalMagicDefense *= other
			totalAccuracy *= other
			totalEvade *= other

		elif (type(other) is Stats):
			totalHP *= other.hp
			totalMP *= other.mp
			totalBallisticAttack *= other.ballisticAttack
			totalMagicAttack *= other.magicAttack
			totalBallisticDefense *= other.ballisticDefense
			totalMagicDefense *= other.magicDefense
			totalAccuracy *= other.accuracy
			totalEvade *= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __truediv__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP /= other
			totalMP /= other
			totalBallisticAttack /= other
			totalMagicAttack /= other
			totalBallisticDefense /= other
			totalMagicDefense /= other
			totalAccuracy /= other
			totalEvade /= other

		elif (type(other) is Stats):
			totalHP /= other.hp
			totalMP /= other.mp
			totalBallisticAttack /= other.ballisticAttack
			totalMagicAttack /= other.magicAttack
			totalBallisticDefense /= other.ballisticDefense
			totalMagicDefense /= other.magicDefense
			totalAccuracy /= other.accuracy
			totalEvade /= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __floordiv__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP //= other
			totalMP //= other
			totalBallisticAttack //= other
			totalMagicAttack //= other
			totalBallisticDefense //= other
			totalMagicDefense //= other
			totalAccuracy //= other
			totalEvade //= other

		elif (type(other) is Stats):
			totalHP //= other.hp
			totalMP //= other.mp
			totalBallisticAttack //= other.ballisticAttack
			totalMagicAttack //= other.magicAttack
			totalBallisticDefense //= other.ballisticDefense
			totalMagicDefense //= other.magicDefense
			totalAccuracy //= other.accuracy
			totalEvade //= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __mod__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP %= other
			totalMP %= other
			totalBallisticAttack %= other
			totalMagicAttack %= other
			totalBallisticDefense %= other
			totalMagicDefense %= other
			totalAccuracy %= other
			totalEvade %= other

		elif (type(other) is Stats):
			totalHP %= other.hp
			totalMP %= other.mp
			totalBallisticAttack %= other.ballisticAttack
			totalMagicAttack %= other.magicAttack
			totalBallisticDefense %= other.ballisticDefense
			totalMagicDefense %= other.magicDefense
			totalAccuracy %= other.accuracy
			totalEvade %= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)

	def __pow__(self, other):
		totalHP = self.hp
		totalMP = self.mp
		totalBallisticAttack = self.ballisticAttack
		totalMagicAttack = self.magicAttack
		totalBallisticDefense = self.ballisticDefense
		totalMagicDefense = self.magicDefense
		totalAccuracy = self.accuracy
		totalEvade = self.evade

		if (type(other) is int):
			totalHP **= other
			totalMP **= other
			totalBallisticAttack **= other
			totalMagicAttack **= other
			totalBallisticDefense **= other
			totalMagicDefense **= other
			totalAccuracy **= other
			totalEvade **= other

		elif (type(other) is Stats):
			totalHP **= other.hp
			totalMP **= other.mp
			totalBallisticAttack **= other.ballisticAttack
			totalMagicAttack **= other.magicAttack
			totalBallisticDefense **= other.ballisticDefense
			totalMagicDefense **= other.magicDefense
			totalAccuracy **= other.accuracy
			totalEvade **= other.evade

		return Stats(
			hp = totalHP,
			mp = totalMP,
			ballisticAttack = totalBallisticAttack,
			magicAttack = totalMagicAttack,
			ballisticDefense = totalBallisticDefense,
			magicDefense = totalMagicDefense,
			accuracy = totalAccuracy,
			evade = totalEvade
		)
	
	def __round__(self, n):
		return Stats(
			hp = int(round(self.hp, n)),
			mp = int(round(self.mp, n)),
			ballisticAttack = int(round(self.ballisticAttack, n)),
			magicAttack = int(round(self.magicAttack, n)),
			ballisticDefense = int(round(self.ballisticDefense, n)),
			magicDefense = int(round(self.magicDefense, n)),
			accuracy = int(round(self.accuracy, n)),
			evade = int(round(self.evade, n))
		)


class StatusEffect:
	def __init__(self, name="Null", effect=Stats(every=0)):
		self.name = name
		self.effect = effect

	def __repr__(self):
		return str(self.name)


class Skill:
	def __init__(self, name="Null", skillType=0, statType=0, element=0, mpChange=0, accuracyMod=0.00, critChance=0.00, randomMod=0.00, targetingTypeA=None, basePowerA=None, buffDebuffA=None, targetingTypeB=None, basePowerB=None, buffDebuffB=None, statusEffect=None, statusEffectDuration=0):
		self.name = textwrap.shorten(str(name), width = 50, placeholder = "")
		self.skillType = skillTypes[skillType]
		self.statType = statTypes[statType]
		self.element = elements[element]
		self.mpChange = mpChange
		self.accuracyMod = accuracyMod
		self.critChance = critChance
		self.randomMod = randomMod
		self.targetingTypeA = targetingTypes[targetingTypeA] if (targetingTypeA != None) else None
		self.basePowerA = basePowerA
		self.buffDebuffA = buffDebuffA
		self.targetingTypeB = targetingTypes[targetingTypeB] if (targetingTypeB != None) else None
		self.basePowerB = basePowerB
		self.buffDebuffB = buffDebuffB

		if (self.skillType == skillTypes[4]):
			pass
		else:
			if (self.targetingTypeA == None or self.basePowerA == None):
				self.targetingTypeA = None
				self.basePowerA = None
				self.buffDebuffA = None
			if (self.targetingTypeB == None or self.basePowerB == None):
				self.targetingTypeB = None
				self.basePowerB = None
				self.buffDebuffB = None
			if ((self.targetingTypeA == None or self.basePowerA == None) and self.targetingTypeB != None and self.basePowerB != None):
				self.targetingTypeA = self.targetingTypeB
				self.basePowerA = self.basePowerB
				self.buffDebuffA = self.buffDebuffB
				self.targetingTypeB = None
				self.basePowerB = None
				self.buffDebuffB = None

		if (self.buffDebuffA != None):
			currentBuffDebuffA = self.buffDebuffA
			allStats = [currentStat for currentStat in dir(currentBuffDebuffA) if not currentStat.startswith('__') and not callable(getattr(currentBuffDebuffA, currentStat))]
			for currentStat in allStats:
				setattr(currentBuffDebuffA, currentStat, 5 * round(getattr(currentBuffDebuffA, currentStat) / 5))

			statsList = [self.buffDebuffA.hp, self.buffDebuffA.mp, self.buffDebuffA.ballisticAttack, self.buffDebuffA.magicAttack, self.buffDebuffA.ballisticDefense, self.buffDebuffA.magicDefense, self.buffDebuffA.accuracy, self.buffDebuffA.evade]
			incorrectBuffsDebuffsAmount = statsList.count(0) < 7 or statsList.count(0) == 0
			if (incorrectBuffsDebuffsAmount):
				highestStat = max(statsList)
				lowestStat = min(statsList)
				if (abs(highestStat) >= abs(lowestStat)):
					chosenStat = highestStat
				else:
					chosenStat = lowestStat
				self.buffDebuffA = Stats(every = chosenStat)

		if (self.buffDebuffB != None):
			currentBuffDebuffB = self.buffDebuffB
			allStats = [currentStat for currentStat in dir(currentBuffDebuffB) if not currentStat.startswith('__') and not callable(getattr(currentBuffDebuffB, currentStat))]
			for currentStat in allStats:
				setattr(currentBuffDebuffB, currentStat, 5 * round(getattr(currentBuffDebuffB, currentStat) / 5))

			statsList = [self.buffDebuffB.hp, self.buffDebuffB.mp, self.buffDebuffB.ballisticAttack, self.buffDebuffB.magicAttack, self.buffDebuffB.ballisticDefense, self.buffDebuffB.magicDefense, self.buffDebuffB.accuracy, self.buffDebuffB.evade]
			incorrectBuffsDebuffsAmount = statsList.count(0) < 7 or statsList.count(0) == 0
			if (incorrectBuffsDebuffsAmount):
				highestStat = max(statsList)
				lowestStat = min(statsList)
				if (abs(highestStat) >= abs(lowestStat)):
					chosenStat = highestStat
				else:
					chosenStat = lowestStat
				self.buffDebuffB = Stats(every = chosenStat)

		if (statusEffect != None):
			self.statusEffect = statusEffectsList[statusEffect]
		else:
			self.statusEffect = None
		self.statusEffectDuration = statusEffectDuration

		self.skillLevel = 1


	def __repr__(self):
		mpChangeString = "{0:+}".format(self.mpChange)
		statTypeString = self.statType if (self.statType != statTypes[0]) else "---"
		elementString = self.element if (self.element != elements[0]) else "---"
		powerAString = "{0}%".format(round(self.basePowerA * 100)) if (self.basePowerA != None) else "---"
		powerBString = "{0}%".format(round(self.basePowerB * 100)) if (self.basePowerB != None) else "---"
		accuracyString = "Guaranteed" if (self.accuracyMod == math.inf) else "{0}%".format(round(self.accuracyMod * 100))
		if (self.buffDebuffA != None):
			if (self.buffDebuffA.hasAllStats):
				effectAString = "All Stats {0:+}%".format(self.buffDebuffA.hp)
			else:
				statsText = ["AC", "BA", "BD", "EV", "HP", "MA", "MD", "MP"]
				allStats = [currentStat for currentStat in dir(self.buffDebuffA) if not currentStat.startswith('__') and not callable(getattr(self.buffDebuffA, currentStat))]
				allStats.remove("hasAllStats")
				for currentStat in allStats:
					if (getattr(self.buffDebuffA, currentStat) != 0):
						effectAString = statsText[allStats.index(currentStat)] + " {0:+}%".format(getattr(self.buffDebuffA, currentStat))
						break
		else:
			effectAString = "None"
		if (self.buffDebuffB != None):
			if (self.buffDebuffB.hasAllStats):
				effectBString = "All Stats {0:+}%".format(self.buffDebuffB.hp)
			else:
				statsText = ["AC", "BA", "BD", "EV", "HP", "MA", "MD", "MP"]
				allStats = [currentStat for currentStat in dir(self.buffDebuffB) if not currentStat.startswith('__') and not callable(getattr(self.buffDebuffB, currentStat))]
				allStats.remove("hasAllStats")
				for currentStat in allStats:
					if (getattr(self.buffDebuffB, currentStat) != 0):
						effectBString = statsText[allStats.index(currentStat)] + " {0:+}%".format(getattr(self.buffDebuffB, currentStat))
						break
		else:
			effectBString = "None"

		if (self.targetingTypeA != None and self.targetingTypeB != None):
			return (
				"+–––––––––––––––––––––––––––––––+"
				+ "\n| " + formatText(self.name, 0, 29) + " |"
				+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
				+ "\n| " + formatText("MP", 0, 5) + " | " + formatText(mpChangeString, 0, 45) + " |"
				+ "\n| " + formatText("Type", 0, 5) + " | " + formatText(statTypeString, 0, 45) + " |"
				+ "\n| " + formatText("Elmnt", 0, 5) + " | " + formatText(elementString, 0, 45) + " |"
				+ "\n| " + formatText("Acc.", 0, 5) + " | " + formatText(accuracyString, 0, 45) + " |"
				+ "\n| " + formatText("Trgt", 0, 5) + " | " + formatText(self.targetingTypeA, 0, 21) + " | " + formatText(self.targetingTypeB, 0, 21) + " |"
				+ "\n| " + formatText("Power", 0, 5) + " | " + formatText(powerAString, 0, 21) + " | " + formatText(powerBString, 0, 21) + " |"
				+ "\n| " + formatText("Effct", 0, 5) + " | " + formatText(effectAString, 0, 21) + " | " + formatText(effectBString, 0, 21) + " |"
				+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
			)
		else:
			return (
				"+–––––––––––––––––––––––––––––––+"
				+ "\n| " + formatText(self.name, 0, 29) + " |"
				+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
				+ "\n| " + formatText("MP", 0, 5) + " | " + formatText(mpChangeString, 0, 45) + " |"
				+ "\n| " + formatText("Type", 0, 5) + " | " + formatText(statTypeString, 0, 45) + " |"
				+ "\n| " + formatText("Elmnt", 0, 5) + " | " + formatText(self.element if (self.element != elements[0]) else "---", 0, 45) + " |"
				+ "\n| " + formatText("Acc.", 0, 5) + " | " + formatText(accuracyString, 0, 45) + " |"
				+ "\n| " + formatText("Trgt", 0, 5) + " | " + formatText(self.targetingTypeA, 0, 45) + " |"
				+ "\n| " + formatText("Power", 0, 5) + " | " + formatText(powerAString, 0, 45) + " |"
				+ "\n| " + formatText("Effct", 0, 5) + " | " + formatText(effectAString, 0, 45) + " |"
				+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
			)


class Item:
	def __init__(self, name="Null", itemType=0, targetingType=0, rawPower=0, relativePower=0, value=0):
		self.name = str(name)
		self.targetingType = targetingTypes[targetingType]
		self.itemType = itemTypes[itemType]
		self.rawPower = rawPower
		self.relativePower = relativePower
		self.value = value

	def __repr__(self):
		if (self.itemType == itemTypes[0]):
			descString = "Heals an ally for " + str(self.rawPower) + " HP or " + str(self.relativePower) + "% HP, whichever is higher."
		elif (self.itemType == itemTypes[1]):
			descString = "Heals an ally for " + str(self.relativePower) + "% MP."
		else:
			descString = "Null"

		descString = textwrap.wrap(descString, width = 45, max_lines = 3, placeholder = "...")
		if (len(descString) < 3):
			temp = len(descString)
			for i in range(3 - temp):
				descString.append("")

		return (
			"+–––––––––––––––––––––––––––––––+"
			+ "\n| " + formatText(self.name, 0, 29) + " |"
			+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
			+ "\n| " + formatText("Desc.", 0, 5) + " | " + formatText(descString[0], 0, 45) + " |"
			+ "\n| " + formatText("", 0, 5) + " | " + formatText(descString[1], 0, 45) + " |"
			+ "\n| " + formatText("", 0, 5) + " | " + formatText(descString[2], 0, 45) + " |"
			+ "\n+–––––––+–––––––––––––––––––––––––––––––––––––––––––––––+"
		)


class PlayerCharacter:
	def __init__(self, name="Null", stats=Stats(), skills=[]):
		self.name = textwrap.shorten(str(name), width = 50, placeholder = "")
		self.curHP = 1
		self.curMP = 1

		self.baseStats = stats

		self.leveledUpStats = Stats(
			hp = round(self.baseStats.hp * (1.05 ** (partyLevel - 1))),
			mp = round(self.baseStats.mp * (1.01 ** (partyLevel - 1))),
			ballisticAttack = round(self.baseStats.ballisticAttack * (1.05 ** (partyLevel - 1))),
			magicAttack = round(self.baseStats.magicAttack * (1.05 ** (partyLevel - 1))),
			ballisticDefense = round(self.baseStats.ballisticDefense * (1.05 ** (partyLevel - 1))),
			magicDefense = round(self.baseStats.magicDefense * (1.05 ** (partyLevel - 1))),
			accuracy = round(self.baseStats.accuracy * (1.05 ** (partyLevel - 1))),
			evade = round(self.baseStats.evade * (1.05 ** (partyLevel - 1)))
		)

		self.skills = skills

		self.buffsDebuffs = Stats(every = 100)

		self.statusEffects = []
		self.statusEffectDurations = []

		self.totalStats = Stats()
		self.evaluateTotalStats()

		self.maxHP = self.totalStats.hp
		self.maxMP = self.totalStats.mp
		self.curHP = self.maxHP
		self.curMP = self.maxMP

	def evaluateTotalStats(self):
		healthRatio = self.curHP / self.totalStats.hp
		manaRatio = self.curMP / self.totalStats.mp

		allStats = [currentStat for currentStat in dir(self.buffsDebuffs) if not currentStat.startswith('__') and not callable(getattr(self.buffsDebuffs, currentStat))]
		for currentStat in allStats:
			setattr(self.buffsDebuffs, currentStat, min(max(getattr(self.buffsDebuffs, currentStat), 50), 200))

		self.totalStats = round(self.leveledUpStats * self.buffsDebuffs / 100, 0)

		self.maxHP = self.totalStats.hp
		self.maxMP = self.totalStats.mp
		self.curHP = int(self.totalStats.hp * healthRatio)
		self.curMP = int(self.totalStats.mp * manaRatio)

	def evaluateCurrentPoints(self):
		self.curHP = min(max(self.curHP, 0), self.maxHP)
		self.curMP = min(max(self.curMP, 0), self.maxMP)


class EnemyCharacter:
	def __init__(self, name="Null", stats=Stats(), skills=[]):
		self.name = name
		# levelConstant = gameRegion * regionBattle
		self.currentLevel = random.choice([max(1, partyLevel - 1), max(1, partyLevel - 1), max(1, partyLevel - 1), partyLevel, partyLevel, partyLevel, partyLevel, min(maxPlayerLevel, partyLevel + 1)])
		self.curHP = 1

		self.baseStats = stats

		self.leveledUpStats = Stats(
			hp = round(self.baseStats.hp * (1.05 ** (self.currentLevel - 1))),
			mp = 1,
			ballisticAttack = round(self.baseStats.ballisticAttack * (1.05 ** (self.currentLevel - 1))),
			magicAttack = round(self.baseStats.magicAttack * (1.05 ** (self.currentLevel - 1))),
			ballisticDefense = round(self.baseStats.ballisticDefense * (1.05 ** (self.currentLevel - 1))),
			magicDefense = round(self.baseStats.magicDefense * (1.05 ** (self.currentLevel - 1))),
			accuracy = round(self.baseStats.accuracy * (1.05 ** (self.currentLevel - 1))),
			evade = round(self.baseStats.evade * (1.05 ** (self.currentLevel - 1)))
		)

		self.skills = skills

		self.buffsDebuffs = Stats(every = 100)

		self.statusEffects = []
		self.statusEffectDurations = []

		self.totalStats = Stats()
		self.evaluateTotalStats()
		# self.checkForLevelUp()

		self.maxHP = self.totalStats.hp
		self.curHP = self.maxHP

	def evaluateTotalStats(self):
		healthRatio = self.curHP / self.totalStats.hp

		self.totalStats = self.leveledUpStats * (self.buffsDebuffs / 100)

		self.maxHP = self.totalStats.hp
		self.curHP = superRound(self.totalStats.hp * healthRatio, int)

	def evaluateCurrentPoints(self):
		self.curHP = min(max(self.curHP, 0), self.maxHP)


	# def checkForLevelUp(self):
	# 	self.baseStats = Stats(
	# 		math.ceil(self.baseStats.hp * (self.levelUpStats.hp ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.ballisticAttack * (self.levelUpStats.ballisticAttack ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.magicAttack * (self.levelUpStats.magicAttack ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.ballisticDefense * (self.levelUpStats.ballisticDefense ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.magicDefense * (self.levelUpStats.magicDefense ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.accuracy * (self.levelUpStats.accuracy ** (self.currentLevel - 1))),
	# 		math.ceil(self.baseStats.evade * (self.levelUpStats.evade ** (self.currentLevel - 1)))
	# 	)
	# 	self.evaluateTotalStats()


# Initialize game data #
maxPlayerLevel = 50

gameRegion = 1
regionBattle = 1
battleTurn = 0

partyMoney = 500
partyLevel = 1
partyCurrentXP = 0
partyNextXP = 200 * (1.15 ** (partyLevel - 1))
partyCurrentFocus = 1.00
partyMaxFocus = 1.00

currentPlayers = [None, None, None]
currentEnemies = [None, None, None]
currentInventory = []

selectedPlayer = -1
selectedEnemy = -1

playersAlive = [None, None, None]
enemiesAlive = [None, None, None]

playersHaveMoved = [None, None, None]
enemiesHaveMoved = [None, None, None]

regions = {
	0: "None",
	1: "Green Grasslands",
	2: "Crude Crimson",
	3: "Dark Depths",
}

skillTypes = {
	0: "–",
	1: "Attack",
	2: "Heal",
	3: "Revive",
	4: "Buff/Debuff",
	5: "Status Effect"
}

statTypes = {
	0: "None",
	1: "Ballistic",
	2: "Magic"
}

targetingTypes = {
	0: "None",
	1: "Self",
	2: "Single Enemy",
	3: "All Enemies",
	4: "Single Ally (Living)",
	5: "All Allies (Living)",
	6: "Single Ally (Dead)",
	7: "All Allies (Dead)",
}

elements = {
	0: "None",
	1: "Non-elemental",
	2: "Fire",
	3: "Water",
	4: "Ice",
	5: "Inherited",
}

resources = {
	0: "None",
	1: " HP",
	2: " MP",
	3: " HP/MP",
	4: "% HP",
	5: "% MP",
	6: "% HP/MP"
}

flatResources = [
	resources[1],
	resources[2],
	resources[3]
]

percentageResources = [
	resources[4],
	resources[5],
	resources[6]
]

itemTypes = {
	0: "HP Recovery",
	1: "MP Recovery"
}

# Contains the list for status effects.
statusEffectsList = {
	0: StatusEffect(),
	1: StatusEffect(
		name = "Defend",
		effect = Stats(ballisticDefense = 50, magicDefense = 50)
	),
	2: StatusEffect(
		name = "Evade",
		effect = Stats(evade = 50)
	)
}

# Contains the list for skills.
skillsList = {
	0: Skill(),
	1: Skill(
		name = "Skip",
		skillType = 0,
		statType = 0,
		element = 0,
		mpChange = 25,
		accuracyMod = 0.00,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 1,
		basePowerA = 0.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	2: Skill(
		name = "Defend",
		skillType = 5,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 1,
		basePowerA = 0.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = 1,
		statusEffectDuration = 3
	),
	3: Skill(
		name = "Evade",
		skillType = 5,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 1,
		basePowerA = 0.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = 2,
		statusEffectDuration = 3
	),
	4: Skill(
		name = "Attack A",
		skillType = 1,
		statType = 1,
		element = 1,
		mpChange = 10,
		accuracyMod = 1.00,
		critChance = 0.10,
		randomMod = 0.10,
		targetingTypeA = 2,
		basePowerA = 1.00,
		buffDebuffA = Stats(hp = -50),
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	5: Skill(
		name = "Attack B",
		skillType = 1,
		statType = 1,
		element = 1,
		mpChange = -10,
		accuracyMod = 1.00,
		critChance = 0.10,
		randomMod = 0.10,
		targetingTypeA = 2,
		basePowerA = 1.50,
		buffDebuffA = Stats(hp = -50),
		targetingTypeB = 3,
		basePowerB = 0.75,
		buffDebuffB = Stats(hp = -20),
		statusEffect = None,
		statusEffectDuration = 0
	),
	6: Skill(
		name = "Attack C",
		skillType = 1,
		statType = 2,
		element = 1,
		mpChange = 10,
		accuracyMod = 1.00,
		critChance = 0.10,
		randomMod = 0.10,
		targetingTypeA = 2,
		basePowerA = 1.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	7: Skill(
		name = "Attack D",
		skillType = 1,
		statType = 2,
		element = 1,
		mpChange = -10,
		accuracyMod = 1.00,
		critChance = 0.10,
		randomMod = 0.10,
		targetingTypeA = 2,
		basePowerA = 1.50,
		buffDebuffA = None,
		targetingTypeB = 3,
		basePowerB = 0.75,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	8: Skill(
		name = "Self-heal",
		skillType = 2,
		statType = 2,
		element = 1,
		mpChange = 10,
		accuracyMod = 1.00,
		critChance = 0.00,
		randomMod = 0.10,
		targetingTypeA = 1,
		basePowerA = 1.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	9: Skill(
		name = "Heal A",
		skillType = 2,
		statType = 2,
		element = 1,
		mpChange = 0,
		accuracyMod = 1.00,
		critChance = 0.00,
		randomMod = 0.10,
		targetingTypeA = 4,
		basePowerA = 1.00,
		buffDebuffA = Stats(every = 50),
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	10: Skill(
		name = "Heal B",
		skillType = 2,
		statType = 2,
		element = 1,
		mpChange = -100,
		accuracyMod = 1.00,
		critChance = 0.00,
		randomMod = 0.10,
		targetingTypeA = 4,
		basePowerA = 1.50,
		buffDebuffA = Stats(every = 50),
		targetingTypeB = 5,
		basePowerB = 0.75,
		buffDebuffB = Stats(every = 20),
		statusEffect = None,
		statusEffectDuration = 0
	),
	11: Skill(
		name = "Revive A",
		skillType = 3,
		statType = 2,
		element = 1,
		mpChange = -25,
		accuracyMod = 1.00,
		critChance = 0.00,
		randomMod = 0.10,
		targetingTypeA = 6,
		basePowerA = 1.00,
		buffDebuffA = None,
		targetingTypeB = None,
		basePowerB = None,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	12: Skill(
		name = "Revive B",
		skillType = 3,
		statType = 2,
		element = 1,
		mpChange = -100,
		accuracyMod = 1.00,
		critChance = 0.00,
		randomMod = 0.10,
		targetingTypeA = 6,
		basePowerA = 1.50,
		buffDebuffA = None,
		targetingTypeB = 7,
		basePowerB = 0.75,
		buffDebuffB = None,
		statusEffect = None,
		statusEffectDuration = 0
	),
	13: Skill(
		name = "Buff A",
		skillType = 4,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 4,
		basePowerA = None,
		buffDebuffA = Stats(ballisticAttack = 50),
		targetingTypeB = 5,
		basePowerB = None,
		buffDebuffB = Stats(ballisticAttack = 20),
		statusEffect = None,
		statusEffectDuration = 0
	),
	14: Skill(
		name = "Buff B",
		skillType = 4,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 4,
		basePowerA = None,
		buffDebuffA = Stats(magicAttack = 50),
		targetingTypeB = 5,
		basePowerB = None,
		buffDebuffB = Stats(magicAttack = 20),
		statusEffect = None,
		statusEffectDuration = 0
	),
	15: Skill(
		name = "Debuff A",
		skillType = 4,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 2,
		basePowerA = None,
		buffDebuffA = Stats(ballisticAttack = -50),
		targetingTypeB = 3,
		basePowerB = None,
		buffDebuffB = Stats(ballisticAttack = -20),
		statusEffect = None,
		statusEffectDuration = 0
	),
	16: Skill(
		name = "Debuff B",
		skillType = 4,
		statType = 0,
		element = 0,
		mpChange = 0,
		accuracyMod = math.inf,
		critChance = 0.00,
		randomMod = 0.00,
		targetingTypeA = 2,
		basePowerA = None,
		buffDebuffA = Stats(magicAttack = -50),
		targetingTypeB = 3,
		basePowerB = None,
		buffDebuffB = Stats(magicAttack = -20),
		statusEffect = None,
		statusEffectDuration = 0
	),
}

# Contains the list for items.
itemsList = {
	0: Item(),
	1: Item(
		name = "HP Recovery Item A",
		itemType = 0,
		targetingType = 4,
		rawPower = 400,
		relativePower = 40,
		value = 150
	),
	2: Item(
		name = "HP Recovery Item B",
		itemType = 0,
		targetingType = 4,
		rawPower = 2000,
		relativePower = 70,
		value = 750
	),
	3: Item(
		name = "HP Recovery Item C",
		itemType = 0,
		targetingType = 4,
		rawPower = 10000,
		relativePower = 100,
		value = 3750
	),
	4: Item(
		name = "HP Recovery Item D",
		itemType = 0,
		targetingType = 5,
		rawPower = 400,
		relativePower = 40,
		value = 150
	),
	5: Item(
		name = "HP Recovery Item E",
		itemType = 0,
		targetingType = 5,
		rawPower = 2000,
		relativePower = 70,
		value = 750
	),
	6: Item(
		name = "HP Recovery Item F",
		itemType = 0,
		targetingType = 5,
		rawPower = 10000,
		relativePower = 100,
		value = 3750
	),
	7: Item(
		name = "MP Recovery Item A",
		itemType = 1,
		targetingType = 4,
		relativePower = 40,
		value = 150
	),
	8: Item(
		name = "MP Recovery Item B",
		itemType = 1,
		targetingType = 4,
		relativePower = 70,
		value = 750
	),
	9: Item(
		name = "MP Recovery Item C",
		itemType = 1,
		targetingType = 4,
		relativePower = 100,
		value = 3750
	),
	10: Item(
		name = "MP Recovery Item D",
		itemType = 1,
		targetingType = 5,
		relativePower = 40,
		value = 150
	),
	11: Item(
		name = "MP Recovery Item E",
		itemType = 1,
		targetingType = 5,
		relativePower = 70,
		value = 750
	),
	12: Item(
		name = "MP Recovery Item F",
		itemType = 1,
		targetingType = 5,
		relativePower = 100,
		value = 3750
	),
}

# Contains the list for players.
playersList = {
	0: PlayerCharacter(),
	1: PlayerCharacter(
		name = "Monkey A",
		stats = Stats(
			hp = 550,
			mp = 100,
			ballisticAttack = 55,
			magicAttack = 50,
			ballisticDefense = 55,
			magicDefense = 50,
			accuracy = 20,
			evade = 20
		),
		skills = [4, 5, 13, 15]
	),
	2: PlayerCharacter(
		name = "Monkey B",
		stats = Stats(
			hp = 450,
			mp = 100,
			ballisticAttack = 50,
			magicAttack = 50,
			ballisticDefense = 50,
			magicDefense = 50,
			accuracy = 25,
			evade = 25
		),
		skills = [6, 7, 14, 16]
	),
	3: PlayerCharacter(
		name = "Monkey C",
		stats = Stats(
			hp = 500,
			mp = 100,
			ballisticAttack = 50,
			magicAttack = 55,
			ballisticDefense = 50,
			magicDefense = 55,
			accuracy = 20,
			evade = 20
		),
		skills = [8, 9, 10, 11, 12]
	),
}

# Contains the list for enemies.
enemiesList = {
	0: EnemyCharacter(),
	1: EnemyCharacter(
		name = "Bloon A",
		stats = Stats(
			hp = 1500,
			mp = 1,
			ballisticAttack = 50,
			magicAttack = 55,
			ballisticDefense = 50,
			magicDefense = 20,
			accuracy = 20,
			evade = 20
		),
		skills = [4, 5]
	),
	2: EnemyCharacter(
		name = "Bloon B",
		stats = Stats(
			hp = 750,
			mp = 1,
			ballisticAttack = 50,
			magicAttack = 50,
			ballisticDefense = 50,
			magicDefense = 25,
			accuracy = 25,
			evade = 25
		),
		skills = [6, 7]
	),
	3: EnemyCharacter(
		name = "Bloon C",
		stats = Stats(
			hp = 1000,
			mp = 1,
			ballisticAttack = 55,
			magicAttack = 50,
			ballisticDefense = 55,
			magicDefense = 20,
			accuracy = 20,
			evade = 20
		),
		skills = [8, 9, 10]
	),
}

# Contains the list for inn actions.
innActions = [
	"Eat",
	"Dine",
	"Rest",
	"Meditate",
	"Nap",
	"Slumber"
]

# Contains the list for inn power modifiers.
innPowers = [
	50,
	100,
	50,
	100,
	50,
	100
]

# Contains the list for inn recovery resources.
innResources = [
	resources[4],
	resources[4],
	resources[5],
	resources[5],
	resources[6],
	resources[6]
]

# Contains the list for costs of inn actions.
innCosts = [
	0,
	0,
	0,
	0,
	0,
	0
]

# Render the main menu of the game. This menu should give the player the option to start a new game, load a saved game, and quit the game.
def renderMainMenu():
	inMainMenu = True

	while (inMainMenu):
		clearScreen()
		print("Title")
		print("–––––")
		print("1. New Game")
		# print("2. Load Game")
		print("9. Debug")
		print("`. Exit")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput == 1):
					renderNewGameMenu()
				# elif (userInput == 2):
					# renderLoadGameMenu()
				elif (userInput == 9):
					renderDebugMenu()
			elif (userInput == "`"):
				inMainMenu = False
		except:
			continue


def renderDebugMenu():
	global skillsList

	currentLocation = "D:\MEGAsync Downloads\python rpg\python rpg.xlsx"
	currentWorkbook = openpyxl.load_workbook(filename = currentLocation, data_only = True)

	allPlayerNames = []
	for index in playersList:
		allPlayerNames.append(playersList[index][0])

	allEnemyNames = []
	for index in enemiesList:
		allEnemyNames.append(enemiesList[index][0])

	choice = -1

	while (choice != 0):
		clearScreen()
		print("Debug")
		print("-----")
		print("1. Update Skills List")
		print("0. Back")

		try:
			choice = int(input("> "))
			if (choice == 1):
				# print(skillsList[0])
				skillsList = {}

				currentSheet = currentWorkbook["SkillsList"]

				currentRow = 2
				while ((currentSheet.cell(currentRow, 2)).value != None):
					currentSkillID = int(currentSheet.cell(currentRow, 1).value)

					currentSkillForEntity = currentSheet.cell(currentRow, 2).value
					if (currentSkillForEntity in allPlayerNames):
						currentSkillForEntity = allPlayerNames.index(currentSkillForEntity)
					elif (currentSkillForEntity in allEnemyNames):
						currentSkillForEntity = allEnemyNames.index(currentSkillForEntity)
					else:
						currentSkillForEntity = 0

					currentSkillName = currentSheet.cell(currentRow, 3).value

					currentSkillMaxLevel = currentSheet.cell(currentRow, 4).value
					if (currentSkillMaxLevel != "–"):
						currentSkillMaxLevel = int(currentSkillMaxLevel)
					else:
						currentSkillMaxLevel = 0

					currentSkillType = currentSheet.cell(currentRow, 5).value
					for i in range(len(skillTypes)):
						if (currentSkillType == skillTypes[i]):
							currentSkillType = i
							break
						if (i == (len(skillTypes) - 1)):
							currentSkillType = 0
							break

					currentSkillStatType = currentSheet.cell(currentRow, 6).value
					for i in range(len(statTypes)):
						if (currentSkillStatType == statTypes[i]):
							currentSkillStatType = i
							break
						if (i == (len(statTypes) - 1)):
							currentSkillStatType = 0
							break

					currentSkillElement = currentSheet.cell(currentRow, 7).value
					for i in range(len(elements)):
						if (currentSkillElement == elements[i]):
							currentSkillElement = i
							break
						if (i == (len(elements) - 1)):
							currentSkillElement = 0
							break

					currentSkillCost = currentSheet.cell(currentRow, 8).value
					if (currentSkillCost != "–"):
						currentSkillCost = int(currentSkillCost)
					else:
						currentSkillCost = 0

					currentTargetingTypeA = currentSheet.cell(currentRow, 9).value
					validTargetingTypeA = False
					for i in range(len(targetingTypes)):
						if (currentTargetingTypeA == targetingTypes[i] and targetingTypes[i] != "–"):
							currentTargetingTypeA = i
							validTargetingTypeA = True
							break
					if (not validTargetingTypeA):
						currentTargetingTypeA = None

					currentTargetingTypeB = currentSheet.cell(currentRow, 10).value
					validTargetingTypeB = False
					for i in range(len(targetingTypes)):
						if (currentTargetingTypeB == targetingTypes[i] and targetingTypes[i] != "–"):
							currentTargetingTypeB = i
							validTargetingTypeB = True
							break
					if (not validTargetingTypeB):
						currentTargetingTypeB = None

					currentBasePowerA = currentSheet.cell(currentRow, 11).value
					if (type(currentBasePowerA) is int or type(currentBasePowerA) is float):
						currentBasePowerA = float("{:.2f}".format(currentBasePowerA))
					else:
						currentBasePowerA = None

					currentBasePowerB = currentSheet.cell(currentRow, 12).value
					if (type(currentBasePowerB) is int or type(currentBasePowerB) is float):
						currentBasePowerB = float("{:.2f}".format(currentBasePowerB))
					else:
						currentBasePowerB = None

					currentAccuracyMod = currentSheet.cell(currentRow, 13).value
					if (currentAccuracyMod != "–"):
						currentAccuracyMod = float("{:.2f}".format(currentAccuracyMod))
					else:
						currentAccuracyMod = 1.00

					currentCritChance = currentSheet.cell(currentRow, 14).value
					if (currentCritChance != "–"):
						currentCritChance = float("{:.2f}".format(currentCritChance))
					else:
						currentCritChance = 0.00

					currentRandomnessMod = currentSheet.cell(currentRow, 15).value
					if (currentRandomnessMod != "–"):
						currentRandomnessMod = float("{:.2f}".format(currentRandomnessMod))
					else:
						currentRandomnessMod = 0.10

					currentStatusEffect = currentSheet.cell(currentRow, 16).value
					if (currentStatusEffect in statusEffectsList):
						currentStatusEffect = str(statusEffectsList.index(currentStatusEffect))
					else:
						currentStatusEffect = None

					currentStatusEffectDuration = currentSheet.cell(currentRow, 17).value
					if (currentStatusEffectDuration != "–"):
						currentStatusEffectDuration = int(currentStatusEffectDuration)
					else:
						currentStatusEffectDuration = 0

					skillsList[currentSkillID] = [
						currentSkillForEntity,
						currentSkillName,
						currentSkillMaxLevel,
						currentSkillType,
						currentSkillStatType,
						currentSkillElement,
						currentSkillCost,
						currentTargetingTypeA,
						currentTargetingTypeB,
						currentBasePowerA,
						currentBasePowerB,
						currentAccuracyMod,
						currentCritChance,
						currentRandomnessMod,
						currentStatusEffect,
						currentStatusEffectDuration
					]

					currentRow += 1

				print(skillsList)
				input()
				choice = 0
		except ValueError:
			continue


# Render the menu for when the player starts a new game. This menu should allow the player to check out and customize their character (such as their name and class). This menu should also allow the player to officially start a new game.
def renderNewGameMenu():
	inNewGameMenu = True
	# selectedCharacters = [0, 0, 0]
	selectedCharacters = [1, 2, 3]

	while (inNewGameMenu):
		clearScreen()
		print("New Game")
		print("--------")
		print("1. " + playersList[selectedCharacters[0]].name) if selectedCharacters[0] != 0 else print("1. Empty Slot")
		print("2. " + playersList[selectedCharacters[1]].name) if selectedCharacters[1] != 0 else print("2. Empty Slot")
		print("3. " + playersList[selectedCharacters[2]].name) if selectedCharacters[2] != 0 else print("3. Empty Slot")
		if (not all(flag == 0 for flag in selectedCharacters)):
			print("Q. Start Game")
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (1 <= userInput <= 3):
					selectedCharacters[userInput - 1] = requestCharacterChange(selectedCharacters[userInput - 1], selectedCharacters)
			elif (userInput.lower() == "q" and not all(flag == 0 for flag in selectedCharacters)):
				for i in range(len(currentPlayers)):
					if (selectedCharacters[i] != 0):
						currentPlayers[i] = copy.deepcopy(playersList[selectedCharacters[i]])
						playersAlive[i] = True
						playersHaveMoved[i] = False
				for i in range(len(itemsList)):
					currentInventory.append(0) if (i == 0) else currentInventory.append(10)
				startGame()
				inNewGameMenu = False
			elif (userInput == "`"):
				inNewGameMenu = False
		except:
			continue


# Render a menu that allows the player to change a character in the party.
def requestCharacterChange(currentCharacter, party):
	changingCharacter = True

	while (changingCharacter):
		clearScreen()
		print("Select Tower")
		print("-------------")
		print("Current Tower in Slot: " + str(playersList[currentCharacter].name))
		for key in playersList:
			if (key != 0):
				print(str(key) + ". " + str(playersList[key].name))
		print("Q. Deselect")
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput != 0 and userInput in playersList):
					changingCharacter = False
			elif (userInput.lower() == "q"):
				userInput = 0
				changingCharacter = False
			elif (userInput == "`"):
				userInput = currentCharacter
				changingCharacter = False
		except:
			continue

	return userInput


# Handles what happens inbetween battles and towns.
def startGame():
	gameRegion = 1
	regionBattle = 1

	while (playersAlive.count(True) > 0):
		if (regionBattle % 11 == 0):
			gameRegion += 1
			regionBattle = 1
			clearScreen()
			proceduralPrint("You entered the " + str(regions[gameRegion % 3]) + ".", "")
		if (regionBattle % 5 == 0 or regionBattle % 5 == 1):
			clearScreen()
			proceduralPrint("You came across a town.", "")
			renderTownActionMenu()
		startBattle()
		regionBattle += 1


# Create an enemy character for the player to fight. Also handles the battle system as well.
def startBattle():
	global selectedPlayer
	global selectedEnemy

	# numOfEnemyCharacters = random.randint(1, 3)
	numOfEnemyCharacters = 3
	for i in range(numOfEnemyCharacters):
		currentEnemies[i] = copy.deepcopy(enemiesList[random.randint(1, 3)])
		enemiesAlive[i] = True
		enemiesHaveMoved[i] = False

	clearScreen()
	proceduralPrint("You encountered a group of enemies!", "")
	progressBattleTurn()

	# While both player characters and enemy characters are alive, run the battle sequence.
	while (playersAlive.count(True) > 0 and enemiesAlive.count(True) > 0):
		# Player turn on odd numbers.
		if (battleTurn % 2 == 1):
			# Select the next available player character.
			while (playersHaveMoved.count(True) != (len(currentPlayers) - currentPlayers.count(None)) and enemiesAlive.count(True) > 0):
				selectedPlayer = 0
				while (selectedPlayer < (len(currentPlayers) - currentPlayers.count(None))):
					if (playersHaveMoved[selectedPlayer] == True or playersHaveMoved[selectedPlayer] == None):
						selectedPlayer += 1
					else:
						break
				renderBattleActionMenu()
				evaluateCharacterStatus()

		# Enemy turn on even numbers.
		elif (battleTurn % 2 == 0):
			# Select the next available enemy character.
			while (enemiesHaveMoved.count(True) != (len(currentEnemies) - currentEnemies.count(None)) and playersAlive.count(True) > 0):
				selectedEnemy = 0
				while (selectedEnemy < (len(currentEnemies) - currentEnemies.count(None))):
					if (enemiesHaveMoved[selectedEnemy] == True or enemiesHaveMoved[selectedEnemy] == None):
						selectedEnemy += 1
					else:
						break
				initiateEnemyAttack()
				evaluateCharacterStatus()
		progressBattleTurn()

	finishBattle()


# Increment the turn in battle by 1.
def progressBattleTurn():
	global battleTurn
	global selectedPlayer
	global selectedEnemy
	global playersHaveMoved
	global enemiesHaveMoved

	battleTurn += 1

	# Enemy-to-player turn transition
	if (battleTurn % 2 == 1):
		selectedPlayer = 0
		selectedEnemy = -1

		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				currentPlayerBuffsDebuffs = currentPlayers[i].buffsDebuffs
				allStats = [currentStat for currentStat in dir(currentPlayerBuffsDebuffs) if not currentStat.startswith('__') and not callable(getattr(currentPlayerBuffsDebuffs, currentStat))]
				for currentStat in allStats:
					if (getattr(currentPlayerBuffsDebuffs, currentStat) > 100):
						setattr(currentPlayerBuffsDebuffs, currentStat, getattr(currentPlayerBuffsDebuffs, currentStat) - min(abs(getattr(currentPlayerBuffsDebuffs, currentStat) - 100), 5))

		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				currentEnemyBuffsDebuffs = currentEnemies[i].buffsDebuffs
				allStats = [currentStat for currentStat in dir(currentEnemyBuffsDebuffs) if not currentStat.startswith('__') and not callable(getattr(currentEnemyBuffsDebuffs, currentStat))]
				for currentStat in allStats:
					if (getattr(currentEnemyBuffsDebuffs, currentStat) > 100):
						setattr(currentEnemyBuffsDebuffs, currentStat, getattr(currentEnemyBuffsDebuffs, currentStat) - min(abs(getattr(currentEnemyBuffsDebuffs, currentStat) - 100), 5))

	# Player-to-enemy turn transition
	elif (battleTurn % 2 == 0):
		selectedPlayer = -1
		selectedEnemy = 0

		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				currentPlayerBuffsDebuffs = currentPlayers[i].buffsDebuffs
				allStats = [currentStat for currentStat in dir(currentPlayerBuffsDebuffs) if not currentStat.startswith('__') and not callable(getattr(currentPlayerBuffsDebuffs, currentStat))]
				for currentStat in allStats:
					if (getattr(currentPlayerBuffsDebuffs, currentStat) < 100):
						setattr(currentPlayerBuffsDebuffs, currentStat, getattr(currentPlayerBuffsDebuffs, currentStat) + min(abs(getattr(currentPlayerBuffsDebuffs, currentStat) - 100), 5))

		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				currentEnemyBuffsDebuffs = currentEnemies[i].buffsDebuffs
				allStats = [currentStat for currentStat in dir(currentEnemyBuffsDebuffs) if not currentStat.startswith('__') and not callable(getattr(currentEnemyBuffsDebuffs, currentStat))]
				for currentStat in allStats:
					if (getattr(currentEnemyBuffsDebuffs, currentStat) < 100):
						setattr(currentEnemyBuffsDebuffs, currentStat, getattr(currentEnemyBuffsDebuffs, currentStat) + min(abs(getattr(currentEnemyBuffsDebuffs, currentStat) - 100), 5))

	for i in range(len(playersHaveMoved)):
		if (playersAlive[i]):
			playersHaveMoved[i] = False
		elif (not playersAlive[i] and playersAlive[i] != None):
			playersHaveMoved[i] = True
	for i in range(len(enemiesHaveMoved)):
		if (enemiesAlive[i]):
			enemiesHaveMoved[i] = False
		elif (not enemiesAlive[i] and enemiesAlive[i] != None):
			enemiesHaveMoved[i] = True
		else:
			continue

	# for i in range(len(currentPlayers)):
	# 	if (currentPlayers[i] != None):
	# 		expiredPlayerStatusEffects = []
	# 		for j in range(len(currentPlayers[i].statusEffects)):
	# 			currentPlayers[i].statusEffectDurations[j] -= 1
	# 			if (currentPlayers[i].statusEffectDurations[j] == 0):
	# 				expiredPlayerStatusEffects.append(j)
	# 		expiredPlayerStatusEffects.reverse()
	# 		for j in expiredPlayerStatusEffects:
	# 			currentPlayers[i].statusEffects.pop(j)
	# 			currentPlayers[i].statusEffectDurations.pop(j)
	# 		currentPlayers[i].evaluateTotalStats()
	# 	else:
	# 		continue

	# for i in range(len(currentEnemies)):
	# 	if (currentEnemies[i] != None):
	# 		expiredEnemyStatusEffects = []
	# 		for j in range(len(currentEnemies[i].statusEffects)):
	# 			currentEnemies[i].statusEffectDurations[j] -= 1
	# 			if (currentEnemies[i].statusEffectDurations[j] == 0):
	# 				expiredEnemyStatusEffects.append(j)
	# 		expiredEnemyStatusEffects.reverse()
	# 		for j in expiredEnemyStatusEffects:
	# 			currentEnemies[i].statusEffects.pop(j)
	# 			currentEnemies[i].statusEffectDurations.pop(j)
	# 		currentEnemies[i].evaluateTotalStats()
	# 	else:
	# 		continue

	evaluateCharacterStatus()


# Determines what happens when a battle is finished.
def finishBattle():
	global partyCurrentXP
	global partyMoney

	clearScreen()

	renderBattleStatusMenu()

	if (playersAlive.count(True) == 0):
		proceduralPrint("\nYour party was killed.", "")
	elif (enemiesAlive.count(True) == 0):
		proceduralPrint("\nYou've killed your enemies.", "")

		gainedXP = 0
		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				gainedXP += round(partyNextXP * (1 + 0.5 * (currentEnemies[i].currentLevel - partyLevel)))
		partyCurrentXP += gainedXP
		proceduralPrint("\nYour party gained " + str(gainedXP) + " XP.", "")

		gainedMoney = 0
		for i in range(len(currentEnemies)):
			if (currentEnemies[i] != None):
				gainedMoney = round(100 * (1.05 ** currentEnemies[i].currentLevel))
		partyMoney += gainedMoney
		proceduralPrint("\nYour party gained $" + str(gainedMoney) + ".", "")

		checkForPartyLevelUp()


# Checks if the party has leveled up. If true, increase each party member's stats.
def checkForPartyLevelUp():
	global partyLevel
	global partyCurrentXP
	global partyNextXP
	global innCosts

	partyLeveledUp = False

	initialPartyLevel = partyLevel
	initialStats = []
	healthRatio = []
	manaRatio = []
	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			initialStats.append(currentPlayers[i].totalStats)
			healthRatio.append(currentPlayers[i].curHP / currentPlayers[i].maxHP)
			manaRatio.append(currentPlayers[i].curMP / currentPlayers[i].maxMP)

	while (partyCurrentXP >= partyNextXP and partyLevel < maxPlayerLevel):
		partyLeveledUp = True
		partyLevel += 1
		partyCurrentXP -= partyNextXP
		partyNextXP *= 1.15
		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				currentPlayers[i].leveledUpStats = Stats(
					hp = round(currentPlayers[i].baseStats.hp * (1.05 ** (partyLevel - 1))),
					mp = round(currentPlayers[i].baseStats.mp * (1.01 ** (partyLevel - 1))),
					ballisticAttack = round(currentPlayers[i].baseStats.ballisticAttack * (1.05 ** (partyLevel - 1))),
					magicAttack = round(currentPlayers[i].baseStats.magicAttack * (1.05 ** (partyLevel - 1))),
					ballisticDefense = round(currentPlayers[i].baseStats.ballisticDefense * (1.05 ** (partyLevel - 1))),
					magicDefense = round(currentPlayers[i].baseStats.magicDefense * (1.05 ** (partyLevel - 1))),
					accuracy = round(currentPlayers[i].baseStats.accuracy * (1.05 ** (partyLevel - 1))),
					evade = round(currentPlayers[i].baseStats.evade * (1.05 ** (partyLevel - 1)))
				)
				currentPlayers[i].evaluateTotalStats()

	innCosts = [
		100 * partyLevel,
		200 * partyLevel,
		100 * partyLevel,
		200 * partyLevel,
		200 * partyLevel,
		400 * partyLevel
	]

	newPartyLevel = partyLevel
	newStats = []
	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			newStats.append(currentPlayers[i].totalStats)

	if (partyLeveledUp):
		changeInStats = []
		for i in range(len(currentPlayers)):
			if (currentPlayers[i] != None):
				changeInStats.append(newStats[i] - initialStats[i])
				currentPlayers[i].curHP = superRound(currentPlayers[i].maxHP * healthRatio[i], int)
				currentPlayers[i].curMP = superRound(currentPlayers[i].maxMP * manaRatio[i], int)

		print("\n+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		print(levelUpUIRow("Level Up! LVL " + str(initialPartyLevel) + " -> LVL " + str(newPartyLevel), 0))
		print("+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		for i in range(3):
			if (currentPlayers[i] != None):
				print(levelUpUIRow(str(currentPlayers[i].name), 0))
				print(levelUpUIRow("", 1))
				print(levelUpUIRow(initialStats[i], 2))
				print(levelUpUIRow(changeInStats[i], 3))
			else:
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))
				print(levelUpUIRow("", 0))

			if (i < 2):
				print("|\t––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––\t|")
		print("+–––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––+")
		proceduralPrint("", "")


# Render a menu that prints out the status of players and enemies when in battle (such as level, HP, and MP).
def renderBattleStatusMenu():
	clearScreen()

	if (partyCurrentFocus >= partyMaxFocus):
		playerFocusBar = "❇  " + superRound(partyCurrentFocus, str) + " />> MAX " + generalUIBar(30, 0, ">", 1, 1)[8:32]
	else:
		playerFocusBar = "❇  " + superRound(partyCurrentFocus, str) + " " + generalUIBar(30, 0, ">", (partyCurrentFocus % 1), 1.00)
	playerProgressText = str(regions[gameRegion % 3]) + " \\ Battle " + str(regionBattle) + " \\ Turn " + str(battleTurn)

	allPlayerTexts = []
	allPlayerBars = []
	allPlayerBuffsDebuffs1 = []
	allPlayerBuffsDebuffs2 = []
	allEnemyTexts = []
	allEnemyBars = []
	allEnemyBuffsDebuffs1 = []
	allEnemyBuffsDebuffs2 = []
	statAttributes = ["hp", "ballisticAttack", "ballisticDefense", "accuracy", "mp", "magicAttack", "magicDefense", "evade"]
	statNames = ["HP", "BA", "BD", "AC", "MP", "MA", "MD", "EV"]
	for i in range(3):
		if (currentPlayers[i] != None):
			currentPlayerName = currentPlayers[i].name
			currentPlayerCurHP = currentPlayers[i].curHP
			currentPlayerCurMP = currentPlayers[i].curMP
			currentPlayerMaxHP = currentPlayers[i].maxHP
			currentPlayerMaxMP = currentPlayers[i].maxMP

			if (len(currentPlayerName) > 14):
				currentPlayerName = currentPlayerName[0:11] + "..."
			# currentPlayerText = "/ " + "{:<14}".format(str(currentPlayerName)) + " / HP " + "{:^4}".format(str(currentPlayerCurHP)) + " MP " + "{:^3}".format(str(currentPlayerCurMP)) + " /"
			currentPlayerText = "/ " + "{:<14}".format(str(currentPlayerName)) + " / ❤  " + "{:^4}".format(str(currentPlayerCurHP)) + " 🗲  " + "{:^3}".format(str(currentPlayerCurMP)) + " /"
			if (i == selectedPlayer):
				currentPlayerText = "> " + currentPlayerText
			else:
				currentPlayerText = "  " + currentPlayerText
			currentPlayerBar = generalUIBar(22, 0, "=", currentPlayerCurHP, currentPlayerMaxHP) + generalUIBar(10, 0, "—", currentPlayerCurMP, currentPlayerMaxMP)
			currentPlayerBuffsDebuffs1 = " | "
			for j in range(4):
				if (getattr(currentPlayers[i].buffsDebuffs, statAttributes[j]) != 100):
					currentPlayerBuffsDebuffs1 += "{0}{1:>4}".format(statNames[j], "{0:+}".format(getattr(currentPlayers[i].buffsDebuffs, statAttributes[j]) - 100))
				else:
					currentPlayerBuffsDebuffs1 += "      "
				if (j < 3):
					currentPlayerBuffsDebuffs1 += "  "
			currentPlayerBuffsDebuffs1 += " |"
			currentPlayerBuffsDebuffs2 = " | "
			for j in range(4, 8):
				if (getattr(currentPlayers[i].buffsDebuffs, statAttributes[j]) != 100):
					currentPlayerBuffsDebuffs2 += "{0}{1:>4}".format(statNames[j], "{0:+}".format(getattr(currentPlayers[i].buffsDebuffs, statAttributes[j]) - 100))
				else:
					currentPlayerBuffsDebuffs2 += "      "
				if (j < 7):
					currentPlayerBuffsDebuffs2 += "  "
			currentPlayerBuffsDebuffs2 += " |"
		else:
			currentPlayerText = ""
			currentPlayerBar = ""
			currentPlayerBuffsDebuffs1 = ""
			currentPlayerBuffsDebuffs2 = ""

		if (currentEnemies[i] != None and enemiesAlive[i]):
			currentEnemyName = "LV " + str(currentEnemies[i].currentLevel) + " " + currentEnemies[i].name
			currentEnemyCurHP = currentEnemies[i].curHP
			currentEnemyMaxHP = currentEnemies[i].maxHP

			if (len(currentEnemyName) > 20):
				currentEnemyName = currentEnemyName[0:17] + "..."
			# currentEnemyText = "\\ HP " + "{:^5}".format(str(currentEnemyCurHP)) + " \\ " + "{:>20}".format(str(currentEnemyName)) + " \\"
			currentEnemyText = "\\ ❤  " + "{:^5}".format(str(currentEnemyCurHP)) + " \\ " + "{:>20}".format(str(currentEnemyName)) + " \\"
			if (i == selectedEnemy):
				currentEnemyText += " <"
			else:
				currentEnemyText += "  "
			currentEnemyBar = generalUIBar(34, 1, "=", currentEnemyCurHP, currentEnemyMaxHP)
			currentEnemyBuffsDebuffs1 = "| "
			for j in range(4):
				if (getattr(currentEnemies[i].buffsDebuffs, statAttributes[j]) != 100):
					currentEnemyBuffsDebuffs1 += "{0}{1:>4}".format(statNames[j], "{0:+}".format(getattr(currentEnemies[i].buffsDebuffs, statAttributes[j]) - 100))
				else:
					currentEnemyBuffsDebuffs1 += "      "
				if (j < 3):
					currentEnemyBuffsDebuffs1 += "  "
			currentEnemyBuffsDebuffs1 += " | "
			currentEnemyBuffsDebuffs2 = "| "
			for j in range(4, 8):
				if (getattr(currentEnemies[i].buffsDebuffs, statAttributes[j]) != 100):
					currentEnemyBuffsDebuffs2 += "{0}{1:>4}".format(statNames[j], "{0:+}".format(getattr(currentEnemies[i].buffsDebuffs, statAttributes[j]) - 100))
				else:
					currentEnemyBuffsDebuffs2 += "      "
				if (j < 7):
					currentEnemyBuffsDebuffs2 += "  "
			currentEnemyBuffsDebuffs2 += " | "
		else:
			currentEnemyText = ""
			currentEnemyBar = ""
			currentEnemyBuffsDebuffs1 = ""
			currentEnemyBuffsDebuffs2 = ""

		allPlayerTexts.append(currentPlayerText)
		allPlayerBars.append(currentPlayerBar)
		allPlayerBuffsDebuffs1.append(currentPlayerBuffsDebuffs1)
		allPlayerBuffsDebuffs2.append(currentPlayerBuffsDebuffs2)
		allEnemyTexts.append(currentEnemyText)
		allEnemyBars.append(currentEnemyBar)
		allEnemyBuffsDebuffs1.append(currentEnemyBuffsDebuffs1)
		allEnemyBuffsDebuffs2.append(currentEnemyBuffsDebuffs2)

	print(battleUIRowDivider())

	print(battleUIRow(playerFocusBar, 0) + battleUIRow(playerProgressText, 1))

	print(battleUIRowDivider())

	for i in range(3):
		print(battleUIRow(allPlayerTexts[i], 0) + battleUIRow(allEnemyTexts[i], 1))
		print(battleUIRow(allPlayerBars[i], 0) + battleUIRow(allEnemyBars[i], 1))
		print(battleUIRow(allPlayerBuffsDebuffs1[i], 0) + battleUIRow(allEnemyBuffsDebuffs1[i], 1))
		print(battleUIRow(allPlayerBuffsDebuffs2[i], 0) + battleUIRow(allEnemyBuffsDebuffs2[i], 1))
		# if (i < 2):
		# 	print("|\t–––––––––––––––––––––––––\t|\t–––––––––––––––––––––––––\t|")

	print(battleUIRowDivider())
	print("|\t|\t|\t|\t|\t|\t|\t|\t|\t|\t|\t|")
	print()


# Render a menu that shows the player their general actions when in battle (such as use a skill or an item).
def renderBattleActionMenu():
	global selectedPlayer

	quickSwitchButtons = ["q", "w", "e"]

	while (playersHaveMoved[selectedPlayer] != True):
		try:
			renderBattleStatusMenu()
		except Exception as e:
			input(e)

		print(currentPlayers[selectedPlayer].name)
		print("-" * len(currentPlayers[selectedPlayer].name))
		print("1. Attack")
		# print("1. Skills")
		print("2. Items")
		print("3. Scan")
		print("8. Defend")
		print("9. Evade")
		for i in range(len(currentPlayers)):
			if (not (selectedPlayer == i) and currentPlayers[i] != None and not playersHaveMoved[i]):
				print(quickSwitchButtons[i].upper() + ". Switch to " + currentPlayers[i].name)
		print("`. " + skillsList[1].name + " (+" + str(skillsList[1].mpChange) + " MP)")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput == 1):
					renderAttackMenu()
				elif (userInput == 2):
					renderItemMenu()
				elif (userInput == 3):
					renderScanMenu()
				elif (userInput == 8):
					castSkill([currentPlayers[selectedPlayer]], currentPlayers[selectedPlayer], skillsList[2], 0)
				elif (userInput == 9):
					castSkill([currentPlayers[selectedPlayer]], currentPlayers[selectedPlayer], skillsList[3], 0)
			elif (userInput.lower() in quickSwitchButtons):
				for i in range(len(currentPlayers)):
					if (userInput.lower() == quickSwitchButtons[i] and not (selectedPlayer == i) and currentPlayers[i] != None and not playersHaveMoved[i]):
						selectedPlayer = i
						break
			elif (userInput == "`"):
				castSkill([currentPlayers[selectedPlayer]], currentPlayers[selectedPlayer], skillsList[1], 0)
		except:
			continue


# Render a menu that shows the player their available skills.
def renderAttackMenu():
	pickingSkill = True

	while (pickingSkill):
		renderBattleStatusMenu()

		print(currentPlayers[selectedPlayer].name + " > Attack")
		print("-" * len(currentPlayers[selectedPlayer].name) + "---------")
		for i in range(len(currentPlayers[selectedPlayer].skills)):
			count = i + 1
			currentSkill = skillsList[currentPlayers[selectedPlayer].skills[i]]

			currentSkillResource = currentPlayers[selectedPlayer].maxMP

			currentSkillMPChange = currentSkill.mpChange

			if (currentSkill == skillsList[0]):
				continue
			else:
				print("{0}. {1} ({2:+} MP)".format(count, currentSkill.name, currentSkillMPChange))
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (1 <= userInput <= len(currentPlayers[selectedPlayer].skills) and currentPlayers[selectedPlayer].skills[userInput - 1] != skillsList[0]):
					userInput -= 1
					chosenSkill = skillsList[currentPlayers[selectedPlayer].skills[userInput]]

					currentPlayerResource = currentPlayers[selectedPlayer].curMP
					chosenSkillResource = currentPlayers[selectedPlayer].maxMP

					chosenSkillMPChange = chosenSkill.mpChange

					if (currentPlayerResource == currentPlayers[selectedPlayer].curMP and currentPlayerResource + chosenSkillMPChange >= 0):
						renderAttackDetails(chosenSkill)
						if (playersHaveMoved[selectedPlayer]):
							pickingSkill = False
			elif (userInput == "`"):
				pickingSkill = False
		except:
			continue


# Render a menu that shows the player more details about the skill they just selected when in battle.
def renderAttackDetails(skill):
	viewingSkillDetails = True

	while (viewingSkillDetails):
		renderBattleStatusMenu()

		print(currentPlayers[selectedPlayer].name + " > Attack > Targeting")
		print("-" * len(currentPlayers[selectedPlayer].name) + "---------------------")
		print(skill)

		hasAvailableTarget = False

		targetAButtons = ["1", "2", "3"]
		if (skill.targetingTypeA == targetingTypes[1]):
			print(targetAButtons[0].upper() + ". " + targetingTypes[1])
			hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[2]):
			for i in range(len(currentEnemies) - enemiesAlive.count(None)):
				if (enemiesAlive[i]):
					print(targetAButtons[i].upper() + ". " + currentEnemies[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[3]):
			print(targetAButtons[0].upper() + ". " + targetingTypes[3])
			hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[4]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (playersAlive[i]):
					print(targetAButtons[i].upper() + ". " + currentPlayers[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[5]):
			print(targetAButtons[0].upper() + ". " + targetingTypes[5])
			hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[6]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (not playersAlive[i]):
					print(targetAButtons[i].upper() + ". " + currentPlayers[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeA == targetingTypes[7]):
			if (playersAlive.count(False) >= 1):
				print(targetAButtons[0].upper() + ". " + targetingTypes[7])
				hasAvailableTarget = True

		targetBButtons = ["q", "w", "e"]
		if (skill.targetingTypeB == targetingTypes[1]):
			print(targetBButtons[0].upper() + ". " + targetingTypes[1])
			hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[2]):
			for i in range(len(currentEnemies) - enemiesAlive.count(None)):
				if (enemiesAlive[i]):
					print(targetBButtons[i].upper() + ". " + currentEnemies[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[3]):
			print(targetBButtons[0].upper() + ". " + targetingTypes[3])
			hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[4]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (playersAlive[i]):
					print(targetBButtons[i].upper() + ". " + currentPlayers[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[5]):
			print(targetBButtons[0].upper() + ". " + targetingTypes[5])
			hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[6]):
			for i in range(len(currentPlayers) - playersAlive.count(None)):
				if (not playersAlive[i]):
					print(targetBButtons[i].upper() + ". " + currentPlayers[i].name)
					hasAvailableTarget = True

		elif (skill.targetingTypeB == targetingTypes[7]):
			if (playersAlive.count(False) >= 1):
				print(targetBButtons[0].upper() + ". " + targetingTypes[7])
				hasAvailableTarget = True

		if (not hasAvailableTarget):
			print("No targets available.")

		print("`. Cancel")

		try:
			userInput = input("> ")
			if (userInput in targetAButtons or userInput in targetBButtons):
				if (userInput in targetAButtons and skill.targetingTypeA != None):
					skillVersion = 0

					if (skill.targetingTypeA == targetingTypes[1]):
						if (userInput == targetAButtons[0]):
							userInput = targetAButtons.index(userInput)
							chosenTarget = [currentPlayers[selectedPlayer]]

					elif (skill.targetingTypeA == targetingTypes[2]):
						userInput = targetAButtons.index(userInput)
						if (enemiesAlive[userInput]):
							chosenTarget = [currentEnemies[userInput]]

					elif (skill.targetingTypeA == targetingTypes[3]):
						if (userInput == targetAButtons[0]):
							userInput = targetAButtons.index(userInput)
							chosenTarget = []
							for i in range(len(currentEnemies)):
								if (enemiesAlive[i] and currentEnemies[i] != None):
									chosenTarget.append(currentEnemies[i])

					elif (skill.targetingTypeA == targetingTypes[4]):
						userInput = targetAButtons.index(userInput)
						if (playersAlive[userInput]):
							chosenTarget = [currentPlayers[userInput]]

					elif (skill.targetingTypeA == targetingTypes[5]):
						if (userInput == targetAButtons[0]):
							userInput = targetAButtons.index(userInput)
							chosenTarget = []
							for i in range(len(currentPlayers)):
								if (playersAlive[i] and currentPlayers[i] != None):
									chosenTarget.append(currentPlayers[i])

					elif (skill.targetingTypeA == targetingTypes[6]):
						userInput = targetAButtons.index(userInput)
						if (not playersAlive[userInput]):
							chosenTarget = [currentPlayers[userInput]]

					elif (skill.targetingTypeA == targetingTypes[7]):
						if (userInput == targetAButtons[0]):
							userInput = targetAButtons.index(userInput)
							if (playersAlive.count(False) >= 1):
								chosenTarget = []
								for i in range(len(currentPlayers)):
									if (not playersAlive[i] and currentPlayers[i] != None):
										chosenTarget.append(currentPlayers[i])

					else:
						chosenTarget = []

				elif (userInput in targetBButtons and skill.targetingTypeB != None):
					skillVersion = 1

					if (skill.targetingTypeB == targetingTypes[1]):
						if (userInput == targetBButtons[0]):
							userInput = targetBButtons.index(userInput)
							chosenTarget = [currentPlayers[selectedPlayer]]

					elif (skill.targetingTypeB == targetingTypes[2]):
						userInput = targetBButtons.index(userInput)
						if (enemiesAlive[userInput]):
							chosenTarget = [currentEnemies[userInput]]

					elif (skill.targetingTypeB == targetingTypes[3]):
						if (userInput == targetBButtons[0]):
							userInput = targetBButtons.index(userInput)
							chosenTarget = []
							for i in range(len(currentEnemies)):
								if (enemiesAlive[i] and currentEnemies[i] != None):
									chosenTarget.append(currentEnemies[i])

					elif (skill.targetingTypeB == targetingTypes[4]):
						userInput = targetBButtons.index(userInput)
						if (playersAlive[userInput]):
							chosenTarget = [currentPlayers[userInput]]

					elif (skill.targetingTypeB == targetingTypes[5]):
						if (userInput == targetBButtons[0]):
							userInput = targetBButtons.index(userInput)
							chosenTarget = []
							for i in range(len(currentPlayers)):
								if (playersAlive[i] and currentPlayers[i] != None):
									chosenTarget.append(currentPlayers[i])

					elif (skill.targetingTypeB == targetingTypes[6]):
						userInput = targetBButtons.index(userInput)
						if (not playersAlive[userInput]):
							chosenTarget = [currentPlayers[userInput]]

					elif (skill.targetingTypeB == targetingTypes[7]):
						if (userInput == targetBButtons[0]):
							userInput = targetBButtons.index(userInput)
							if (playersAlive.count(False) >= 1):
								chosenTarget = []
								for i in range(len(currentPlayers)):
									if (not playersAlive[i] and currentPlayers[i] != None):
										chosenTarget.append(currentPlayers[i])

					else:
						chosenTarget = []

				castSkill(chosenTarget, currentPlayers[selectedPlayer], skill, skillVersion)
				viewingSkillDetails = False

			elif (userInput == "`"):
				viewingSkillDetails = False
		except:
			continue


# Render a menu that shows the player all the items.
def renderItemMenu():
	pickingItem = True

	while (pickingItem):
		renderBattleStatusMenu()

		print(currentPlayers[selectedPlayer].name + " > Items")
		print("-" * len(currentPlayers[selectedPlayer].name) + "--------")
		for i in range(1, len(itemsList)):
			print("{0}. {1} ({2})".format(i, itemsList[i].name, currentInventory[i]))
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (1 <= userInput <= len(itemsList)):
					renderItemDetails(userInput)
					if (playersHaveMoved[selectedPlayer]):
						pickingItem = False
			elif (userInput == "`"):
				pickingItem = False
		except:
			continue


# Render a menu that shows the player more details about the item they just selected when in battle.
def renderItemDetails(itemIndex):
	viewingItemDetails = True
	chosenItem = itemsList[itemIndex]

	while (viewingItemDetails):
		renderBattleStatusMenu()

		print(currentPlayers[selectedPlayer].name + " > Items > Use")
		print("-" * len(currentPlayers[selectedPlayer].name) + "--------------")
		print(chosenItem)

		hasAvailableTarget = False

		if (currentInventory[itemIndex] > 0):
			if (chosenItem.targetingType == targetingTypes[4]):
				for i in range(len(currentPlayers) - playersAlive.count(None)):
					if (playersAlive[i]):
						print(str(i + 1) + ". " + currentPlayers[i].name)
						hasAvailableTarget = True

			if (chosenItem.targetingType == targetingTypes[5]):
				print("1. " + targetingTypes[5])
				hasAvailableTarget = True


			if (not hasAvailableTarget):
				print("No targets available.")
		else:
			print("No " + chosenItem.name + "s in inventory.")

		print("`. Cancel")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (chosenItem.targetingType == targetingTypes[4]):
					if (playersAlive[userInput - 1]):
						chosenTarget = [currentPlayers[userInput - 1]]
				elif (chosenItem.targetingType == targetingTypes[5]):
					if (userInput == 1):
						chosenTarget = []
						for i in range(len(currentPlayers)):
							if (playersAlive[i] and currentPlayers[i] != None):
								chosenTarget.append(currentPlayers[i])
				else:
					chosenTarget = []

				useItem(chosenTarget, currentPlayers[selectedPlayer], itemIndex)
				viewingItemDetails = False
			elif (userInput == "`"):
				viewingItemDetails = False
		except:
			continue


# Render a menu that shows the player who they can scan.
def renderScanMenu():
	scanningCharacters = True

	while (scanningCharacters):
		renderBattleStatusMenu()
		print(currentPlayers[selectedPlayer].name + " > Scan")
		print("-" * len(currentPlayers[selectedPlayer].name) + "-------")
		print("1. " + str(currentPlayers[0].name))
		print("2. " + str(currentEnemies[0].name))
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput == 1):
					renderPlayerScanMenu()
				elif (userInput == 2):
					print(currentEnemies[0])
			elif (userInput == "`"):
				scanningCharacters = False
		except:
			continue


# Render a menu that allows the player to scan themselves.
def renderPlayerScanMenu():
	scanningPlayerCharacter = True

	while (scanningPlayerCharacter):
		renderBattleStatusMenu()

		print(currentPlayers[selectedPlayer].name + " > Scan > " + currentPlayers[selectedPlayer].name)
		print("-" * len(currentPlayers[selectedPlayer].name) + "----------" + "-" * len(currentPlayers[selectedPlayer].name))
		print("1. Base Stats")
		print("2. Status Effects")
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput == 1):
					proceduralPrint("\nBase Stats\n" + str(currentPlayers[0].baseStats), "\n")
					scanningPlayerCharacter = False
				elif (userInput == 2):
					statusEffectsString = ""
					if (len(currentPlayers[0].statusEffects) > 0):
						for i in range(len(currentPlayers[0].statusEffects)):
							statusEffectsString += "\n" + str(currentPlayers[0].statusEffects[i]) + " (" + str(currentPlayers[0].statusEffectDurations[i]) + ")"
							# statusEffectsString += str(currentPlayers[0].statusEffects[i])
							# statusEffectsString += " ("
							# statusEffectsString += str(currentPlayers[0].statusEffectDurations[i])
							# statusEffectsString += ")"
					else:
						statusEffectsString = "\nN/A"

					proceduralPrint(statusEffectsString, "\n")
					scanningPlayerCharacter = False
		except:
			continue


# Have the enemy perform an action, such as casting a skill or using an item.
def initiateEnemyAttack():
	global currentPlayers
	global currentEnemies

	renderBattleStatusMenu()

	possibleSkills = []
	for i in range(len(currentEnemies[selectedEnemy].skills)):
		possibleSkill = skillsList[currentEnemies[selectedEnemy].skills[i]]
		if (possibleSkill.name != "Empty" and possibleSkill.name != "Struggle"):
			possibleSkills.append(possibleSkill)

	# if (len(possibleSkills) == 0):
	# 	for currentKey in skillsList:
	# 		if (skillsList[currentKey].name == "Struggle"):
	# 			possibleSkills.append(str(skillsList[currentKey]))
	# 			break

	randomSkill = random.choice(possibleSkills)
	if (randomSkill.targetingTypeA == targetingTypes[1]):
		castSkill([currentEnemies[selectedEnemy]], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[2]):
		availableTargets = []
		for i in range(len(playersAlive)):
			if (playersAlive[i]):
				availableTargets.append(currentPlayers[i])
		randomTarget = random.choice(availableTargets)
		castSkill([randomTarget], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[3]):
		availableTargets = []
		for i in range(len(currentPlayers)):
			if (playersAlive[i] and currentPlayers[i] != None):
				availableTargets.append(currentPlayers[i])
		castSkill(availableTargets, currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[4]):
		availableTargets = []
		for i in range(len(enemiesAlive)):
			if (enemiesAlive[i]):
				availableTargets.append(currentEnemies[i])
		randomTarget = random.choice(availableTargets)
		castSkill([randomTarget], currentEnemies[selectedEnemy], randomSkill, 0)

	elif (randomSkill.targetingTypeA == targetingTypes[5]):
		availableTargets = []
		for i in range(len(currentEnemies)):
			if (enemiesAlive[i] and currentEnemies[i] != None):
				availableTargets.append(currentEnemies[i])
		castSkill(availableTargets, currentEnemies[selectedEnemy], randomSkill, 0)

	else:
		castSkill([currentEnemies[selectedEnemy]], currentEnemies[selectedEnemy], randomSkill, 0)


# Check the status of all players and enemies.
def evaluateCharacterStatus():
	global currentPlayers
	global currentEnemies
	global playersAlive
	global enemiesAlive

	for i in range(len(currentPlayers)):
		if (currentPlayers[i] != None):
			playersAlive[i] = currentPlayers[i].curHP > 0
	for i in range(len(currentEnemies)):
		if (currentEnemies[i] != None):
			enemiesAlive[i] = currentEnemies[i].curHP > 0


# Render a menu that prints out the status of the player when in town (such as level, HP, MP, and money).
def renderTownStatusMenu():
	clearScreen()
	print("+–––––––––––––––––––––––––––––––––––––––+")
	print("| " + str(regions[gameRegion % 3]) + " | Town")
	print("+–––––––––––––––––––––––––––––––––––––––+")
	print("| " + str(currentPlayers[0].name) + " | LV " + str(partyLevel))
	print("| HP | " + str(currentPlayers[0].curHP) + "/" + str(currentPlayers[0].maxHP))
	print("| MP | " + str(currentPlayers[0].curMP) + "/" + str(currentPlayers[0].maxMP))
	print("| $$ | " + str(partyMoney))
	print("+–––––––––––––––––––––––––––––––––––––––+")


# Render a menu that shows the player their general actions when in town (go to the shop or enter the inn).
def renderTownActionMenu():
	inTownMenu = True

	while (inTownMenu):
		renderTownStatusMenu()

		print("\nTown")
		print("----")
		print("1. Shop")
		print("2. Inn")
		# print("9. Save Game")
		print("`. Leave")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				if (userInput == 1):
					renderShopMenu()
				elif (userInput == 2):
					renderInnMenu()
			elif (userInput == "`"):
				inTownMenu = False
		except:
			continue


# Render a menu that shows the player the shop in town.
def renderShopMenu():
	inShop = True

	while (inShop):
		renderTownStatusMenu()

		print("\nTown > Shop")
		print("-----------")
		for i in range(len(currentPlayers[0].itemInventory)):
			currentItemAmount = currentPlayers[0].itemInventory[i]
			print(str(i + 1) + ". ", end="")
			print(itemsList[i].name + " (" + str(currentItemAmount) + ") ($" + str(itemsList[i].value) + ")")
		print("`. Leave")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (1 <= userInput <= len(itemsList)):
					userInput -= 1
					chosenItemAmount = currentPlayers[0].itemInventory[userInput]

					if (chosenItemAmount > 0):
						renderShopDetails(userInput)
			elif (userInput == "`"):
				inShop = False
		except:
			continue


# Render a menu that shows the player more details about the item they just selected when in town.
def renderShopDetails(itemIndex):
	viewingItemDetails = True
	chosenItem = itemsList[itemIndex]

	while (viewingItemDetails):
		renderTownStatusMenu()

		print("Town > Shop > Item Details")
		print("--------------------------")
		print(chosenItem)
		print(">0. Buy current item at 100% value")
		print("<0. Sell current item at 20% value")
		print("`. Back")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (userInput > 0):
					totalCost = chosenItem.value * userInput
					if (partyMoney >= totalCost):
						partyMoney -= totalCost
						currentPlayers[0].itemInventory[itemIndex] += userInput
						if (userInput == 1):
							proceduralPrint("\n" + str(currentPlayers[0].name) + " spent $" + str(totalCost) + " to buy 1 " + str(chosenItem.name) + ".", "")
						else:
							proceduralPrint("\n" + str(currentPlayers[0].name) + " spent $" + str(totalCost) + " to buy " + str(userInput) + " " + str(chosenItem.name) + "s.", "")
						viewingItemDetails = False
				elif (userInput < 0):
					userInput = abs(userInput)
					if (currentPlayers[0].itemInventory[itemIndex] >= userInput):
						totalProfit = (chosenItem.value * (20/100)) * userInput
						totalProfit = superRound(totalProfit, int)
						partyMoney += totalProfit
						currentPlayers[0].itemInventory[itemIndex] -= userInput
						if (userInput == 1):
							proceduralPrint("\n" + str(currentPlayers[0].name) + " sold 1 " + str(chosenItem.name) + " for $" + str(totalProfit) + ".", "")
						else:
							proceduralPrint("\n" + str(currentPlayers[0].name) + " sold " + str(userInput) + " " + str(chosenItem.name) + "s for $" + str(totalProfit) + ".", "")
						viewingItemDetails = False
			elif (userInput == "`"):
				viewingItemDetails = False
		except:
			continue


# Render a menu that shows the player the inn in town.
def renderInnMenu():
	inInn = True

	while (inInn):
		renderTownStatusMenu()

		print("Town > Inn")
		print("----------")
		print("1. " + str(innActions[0]) + " (50% HP) ($" + str(innCosts[0]) + ")")
		print("2. " + str(innActions[1]) + " (100% HP) ($" + str(innCosts[1]) + ")")
		print("3. " + str(innActions[2]) + " (50% MP) ($" + str(innCosts[2]) + ")")
		print("4. " + str(innActions[3]) + " (100% MP) ($" + str(innCosts[3]) + ")")
		print("5. " + str(innActions[4]) + " (50% HP/MP) ($" + str(innCosts[4]) + ")")
		print("6. " + str(innActions[5]) + " (100% HP/MP) ($" + str(innCosts[5]) + ")")
		print("`. Leave")

		try:
			userInput = input("> ")
			if (userInput.isdecimal()):
				userInput = int(userInput)
				if (1 <= userInput <= 6):
					userInput -= 1
					chosenInnCosts = innCosts[userInput]

					if (partyMoney >= chosenInnCosts):
						renderInnDetails(userInput)
						inInn = False
			elif (userInput == "`"):
				inInn = False
		except:
			continue


# Render a menu that shows the player more details about the inn action they just selected when in town.
def renderInnDetails(action):
	viewingActionDetails = True

	while (viewingActionDetails):
		renderTownStatusMenu()
		print("Town > Inn > Action Details")
		print("---------------------------")
		print("+–––––––+–––––––––––––––––––––––––––––––+")
		print("|Action\t| " + str(innActions[action]))
		print("|Power\t| " + str(innPowers[action]) + str(innResources[action]))
		print("|Cost\t| " + str(innCosts[action]))
		print("+–––––––+–––––––––––––––––––––––––––––––+")
		print("\n1. Confirm")
		print("`. Cancel")

		try:
			userInput = int(input("> "))
			if (userInput.isdecimal()):
				if (userInput == 1):
					chosenInnPower = innPowers[action]
					chosenInnResource = innResources[action]
					chosenInnCosts = innCosts[action]

					partyMoney -= chosenInnCosts

					if (chosenInnResource in flatResources):
						currentHPRecovery = superRound(chosenInnPower, int)
						currentMPRecovery = superRound(chosenInnPower, int)
					elif (chosenInnResource in percentageResources):
						currentHPRecovery = currentPlayers[0].maxHP * (chosenInnPower / 100)
						currentMPRecovery = currentPlayers[0].maxMP * (chosenInnPower / 100)
						currentHPRecovery = superRound(currentHPRecovery, int)
						currentMPRecovery = superRound(currentMPRecovery, int)
					else:
						currentHPRecovery = superRound(chosenInnPower, int)
						currentMPRecovery = superRound(chosenInnPower, int)

					if (chosenInnResource == resources[1] or chosenInnResource == resources[4]):
						currentPlayers[0].curHP += currentHPRecovery
						currentPlayers[0].evaluateCurrentPoints()

						proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentHPRecovery) + " HP.", "")
						viewingActionDetails = False
					elif (chosenInnResource == resources[2] or chosenInnResource == resources[5]):
						currentPlayers[0].curMP += currentMPRecovery
						currentPlayers[0].evaluateCurrentPoints()

						proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentMPRecovery) + " MP.", "")
						viewingActionDetails = False
					elif (chosenInnResource == resources[3] or chosenInnResource == resources[6]):
						currentPlayers[0].curHP += currentHPRecovery
						currentPlayers[0].curMP += currentMPRecovery
						currentPlayers[0].evaluateCurrentPoints()

						proceduralPrint("\n" + str(currentPlayers[0].name) + " has recovered " + str(currentHPRecovery) + " HP and " + str(currentMPRecovery) + " MP.", "")
						viewingActionDetails = False
			elif (userInput == "`"):
				viewingActionDetails = False
		except:
			continue


# The user applies a specified status effect onto the target for a specified amount of turns.
def applyStatusEffect(target, user, statusEffect, statusEffectDuration):
	for i in range(len(target)):
		# If the target already has the status effect, reset the duration.
		if (statusEffect in target[i].statusEffects):
			statusEffectIndex = target[i].statusEffects.index(statusEffect)
			target[i].statusEffectDurations[statusEffectIndex] = statusEffectDuration + 1

		# Otherwise, the status effect is applied onto the target.
		else:
			target[i].statusEffects.append(statusEffect)
			target[i].statusEffectDurations.append(statusEffectDuration + 1)
		target[i].evaluateTotalStats()

		proceduralPrint(str(target[i].name) + " received " + str(statusEffect.name) + ".", "")


# The user casts a specified skill onto the target(s).
def castSkill(targets, user, skill, version):
	global battleTurn
	global playersHaveMoved
	global enemiesHaveMoved

	clearScreen()
	renderBattleStatusMenu()

	# Get user text for strings
	userName = user.name
	targetNames = []
	for i in range(len(targets)):
		targetNames.append(targets[i].name)

	if (type(user) is PlayerCharacter):
		user.curMP += skill.mpChange

	if (version == 1):
		skillPower = skill.basePowerB
		skillBuffDebuff = skill.buffDebuffB
	else:
		skillPower = skill.basePowerA
		skillBuffDebuff = skill.buffDebuffA

	userAccuracyStat = user.totalStats.accuracy
	targetEvadeStat = []
	for i in range(len(targets)):
		targetEvadeStat.append(targets[i].totalStats.evade)

	proceduralPrint("\n" + userName + " casted " + skill.name + ".", "")

	if (skill.skillType == skillTypes[1]):
		damageMod = 1

		userAttackStat = None
		targetDefenseStat = []

		if (skill.statType == statTypes[1]):
			userAttackStat = user.totalStats.ballisticAttack
			for i in range(len(targets)):
				targetDefenseStat.append(targets[i].totalStats.ballisticDefense)
		elif (skill.statType == statTypes[2]):
			userAttackStat = user.totalStats.magicAttack
			for i in range(len(targets)):
				targetDefenseStat.append(targets[i].totalStats.magicDefense)
		else:
			userAttackStat = 1
			for i in range(len(targets)):
				targetDefenseStat.append(1)

		allPotDamage = []
		allHitRequirements = []
		allCritRequirements = []
		for i in range(len(targets)):
			currentPotDamage = math.ceil(4 * userAttackStat ** 2 / (userAttackStat + targetDefenseStat[i]) * skillPower)
			if (currentPotDamage <= 0):
				currentPotDamage = 1
			allPotDamage.append(currentPotDamage)

			currentHitRequirement = random.randint(0, 99)
			allHitRequirements.append(currentHitRequirement)

			currentCritRequirement = random.randint(0, 99)
			allCritRequirements.append(currentCritRequirement)

		allRegAccuracy = []
		for i in range(len(targets)):
			allRegAccuracy.append(userAccuracyStat / targetEvadeStat[i] * skill.accuracyMod * 100)

		allCritAccuracy = []
		for i in range(len(targets)):
			allCritAccuracy.append(skill.critChance * 100)

		allDamageDealt = []
		affectedTargets = []
		missedTargets = []
		for i in range(len(targets)):
			currentDamage = round(allPotDamage[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			if (allRegAccuracy[i] > allHitRequirements[i]):
				if (allCritAccuracy[i] > allCritRequirements[i]):
					currentDamage = round(currentDamage * 1.5)
					criticalString = " critical"
				else:
					criticalString = ""

				targets[i].curHP -= currentDamage

				allDamageDealt.append(currentDamage)
			else:
				allDamageDealt.append(None)

		currentEventText = ""
		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			if (allDamageDealt[i] != None):
				currentEventText += "{0} dealt {1}{2} damage to {3}".format(currentUserName, allDamageDealt[i], criticalString, targetNames[i])
			else:
				currentEventText += "{0} missed their attack on {1}".format(currentUserName, targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."
		scrollingPrint(currentEventText)

		if (skillBuffDebuff != None):
			currentEventText = ""
			count = 0

			if (skillBuffDebuff.hasAllStats):
				buffDebuffActionString = "buffed" if (skillBuffDebuff.hp >= 0) else "debuffed"
				buffDebuffStatString = "All Stats"
				buffDebuffStrengthString = "{0}%".format(abs(skillBuffDebuff.hp))
			else:
				statsText = ["Accuracy", "Ballistic Attack", "Ballistic Defense", "Evade", "HP", "Magic Attack", "Magic Defense", "MP"]
				allStats = [currentStat for currentStat in dir(skillBuffDebuff) if not currentStat.startswith('__') and not callable(getattr(skillBuffDebuff, currentStat))]
				allStats.remove("hasAllStats")
				for currentStat in allStats:
					if (getattr(skillBuffDebuff, currentStat) != 0):
						buffDebuffActionString = "buffed" if (getattr(skillBuffDebuff, currentStat) >= 0) else "debuffed"
						buffDebuffStatString = statsText[allStats.index(currentStat)]
						buffDebuffStrengthString = "{0}%".format(abs(getattr(skillBuffDebuff, currentStat)))
						break

			for i in range(len(targets)):
				if (allDamageDealt[i] != None):
					lastDamageDealt = i

			for i in range(len(targets)):
				if (allDamageDealt[i] != None):
					targets[i].buffsDebuffs += skillBuffDebuff

					currentUserName = userName if (count == 0) else formatText("and ", 2, len(userName)) if (i == lastDamageDealt) else " " * len(userName)

					currentEventText += "{0} {1} {2}'s {3} by {4}".format(currentUserName, buffDebuffActionString, targetNames[i], buffDebuffStatString, buffDebuffStrengthString)

					if (i < lastDamageDealt):
						currentEventText += ",\n"
					else:
						currentEventText += "."

					count += 1
			scrollingPrint(currentEventText)

		if (skill.statusEffect != None):
			for i in range(len(targets)):
				applyStatusEffect(targets[i], user, skill.statusEffect, skill.statusEffectDuration)

	elif (skill.skillType == skillTypes[2]):
		if (skill.statType == statTypes[1]):
			userHealStat = user.totalStats.ballisticAttack
		elif (skill.statType == statTypes[2]):
			userHealStat = user.totalStats.magicAttack
		else:
			userHealStat = 1

		allPotHealing = []
		for i in range(len(targets)):
			currentHealing = math.ceil(userHealStat ** 2 / userHealStat * skillPower)
			if (type(targets[0]) is PlayerCharacter):
				currentHealing *= 4
			else:
				currentHealing *= 1.5
			if currentHealing < 0:
				currentHealing = 1
			allPotHealing.append(currentHealing)

		allHealingDealt = []
		for i in range(len(targets)):
			currentHealing = round(allPotHealing[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			targets[i].curHP += currentHealing

			allHealingDealt.append(currentHealing)

		currentEventText = ""
		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			if (allHealingDealt[i] != None):
				currentEventText += "{0} dealt {1} healing to {2}".format(currentUserName, allHealingDealt[i], targetNames[i])
			else:
				currentEventText += "{0} missed their heal on {1}".format(currentUserName, targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."
		scrollingPrint(currentEventText)

		if (skillBuffDebuff != None):
			currentEventText = ""
			count = 0

			if (skillBuffDebuff.hasAllStats):
				buffDebuffActionString = "buffed" if (skillBuffDebuff.hp >= 0) else "debuffed"
				buffDebuffStatString = "All Stats"
				buffDebuffStrengthString = "{0}%".format(abs(skillBuffDebuff.hp))
			else:
				statsText = ["Accuracy", "Ballistic Attack", "Ballistic Defense", "Evade", "HP", "Magic Attack", "Magic Defense", "MP"]
				allStats = [currentStat for currentStat in dir(skillBuffDebuff) if not currentStat.startswith('__') and not callable(getattr(skillBuffDebuff, currentStat))]
				allStats.remove("hasAllStats")
				for currentStat in allStats:
					if (getattr(skillBuffDebuff, currentStat) != 0):
						buffDebuffActionString = "buffed" if (getattr(skillBuffDebuff, currentStat) >= 0) else "debuffed"
						buffDebuffStatString = statsText[allStats.index(currentStat)]
						buffDebuffStrengthString = "{0}%".format(abs(getattr(skillBuffDebuff, currentStat)))
						break

			for i in range(len(targets)):
				if (allHealingDealt[i] != None):
					lastDamageDealt = i

			for i in range(len(targets)):
				if (allHealingDealt[i] != None):
					targets[i].buffsDebuffs += skillBuffDebuff

					currentUserName = userName if (count == 0) else formatText("and ", 2, len(userName)) if (i == lastDamageDealt) else " " * len(userName)

					currentEventText += "{0} {1} {2}'s {3} by {4}".format(currentUserName, buffDebuffActionString, targetNames[i], buffDebuffStatString, buffDebuffStrengthString)

					if (i < lastDamageDealt):
						currentEventText += ",\n"
					else:
						currentEventText += "."

					count += 1
			scrollingPrint(currentEventText)

		if (skill.statusEffect != None):
			for i in range(len(targets)):
				applyStatusEffect(targets[i], user, skill.statusEffect, skill.statusEffectDuration)

	elif (skill.skillType == skillTypes[3]):
		for i in range(len(targets)):
			if (targets[i].curHP > 0):
				targets.pop(targets[i])
				targetNames.pop(targetNames[i])
				targetEvadeStat.pop(targetEvadeStat[i])

		if (skill.statType == statTypes[1]):
			userHealStat = user.totalStats.ballisticAttack
		elif (skill.statType == statTypes[2]):
			userHealStat = user.totalStats.magicAttack
		else:
			userHealStat = 1

		allPotHealing = []
		for i in range(len(targets)):
			currentHealing = round(((userHealStat ** 2) / (userHealStat)) * skillPower)
			if currentHealing < 0:
				currentHealing = 0
			allPotHealing.append(currentHealing)

		allHealingDealt = []
		for i in range(len(targets)):
			currentHealing = round(allPotHealing[i] * random.uniform((1.00 - (skill.randomMod / 2)), (1.00 + (skill.randomMod / 2))))

			targets[i].curHP += currentHealing
			if (user in currentPlayers):
				playersHaveMoved[currentPlayers.index(targets[i])] = False
			elif (user in currentEnemies):
				enemiesHaveMoved[currentEnemies.index(targets[i])] = False

			allHealingDealt.append(currentHealing)

		currentEventText = ""
		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			if (allHealingDealt[i] != None):
				currentEventText += "{0} revived {1} with {2} health".format(currentUserName, targetNames[i], allHealingDealt[i])
			else:
				currentEventText += "{0} missed their revival on {1}".format(currentUserName, targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."
		scrollingPrint(currentEventText)

	elif (skill.skillType == skillTypes[4]):
		allHitRequirements = []
		for i in range(len(targets)):
			currentHitRequirement = random.randint(0, 99)
			allHitRequirements.append(currentHitRequirement)

		allRegAccuracy = []
		for i in range(len(targets)):
			allRegAccuracy.append(userAccuracyStat / targetEvadeStat[i] * skill.accuracyMod * 100)

		for i in range(len(targets)):
			if (allRegAccuracy[i] >= allHitRequirements[i]):
				targets[i].buffsDebuffs += skillBuffDebuff

		if (skillBuffDebuff.hasAllStats):
			buffDebuffActionString = "buffed" if (skillBuffDebuff.hp >= 0) else "debuffed"
			buffDebuffStatString = "All Stats"
			buffDebuffStrengthString = "{0}%".format(abs(skillBuffDebuff.hp))
		else:
			statsText = ["Accuracy", "Ballistic Attack", "Ballistic Defense", "Evade", "HP", "Magic Attack", "Magic Defense", "MP"]
			allStats = [currentStat for currentStat in dir(skillBuffDebuff) if not currentStat.startswith('__') and not callable(getattr(skillBuffDebuff, currentStat))]
			allStats.remove("hasAllStats")
			for currentStat in allStats:
				if (getattr(skillBuffDebuff, currentStat) != 0):
					buffDebuffActionString = "buffed" if (getattr(skillBuffDebuff, currentStat) >= 0) else "debuffed"
					buffDebuffStatString = statsText[allStats.index(currentStat)]
					buffDebuffStrengthString = "{0}%".format(abs(getattr(skillBuffDebuff, currentStat)))
					break

		currentEventText = ""
		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			if (allRegAccuracy[i] >= allHitRequirements[i]):
				currentEventText += "{0} {1} {2}'s {3} by {4}".format(currentUserName, buffDebuffActionString, targetNames[i], buffDebuffStatString, buffDebuffStrengthString)
			else:
				currentEventText += "{0} missed their {1} on {2}".format(currentUserName, buffDebuffActionString, targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."
		scrollingPrint(currentEventText)

	elif (skill.skillType == skillTypes[5]):
		applyStatusEffect(targets, user, skill.statusEffect, skill.statusEffectDuration)

	user.evaluateTotalStats()
	user.evaluateCurrentPoints()
	for i in range(len(targets)):
		targets[i].evaluateTotalStats()
		targets[i].evaluateCurrentPoints()

	if (user in currentPlayers):
		playersHaveMoved[selectedPlayer] = True
	elif (user in currentEnemies):
		enemiesHaveMoved[selectedEnemy] = True


# The user uses an item onto the target(s).
def useItem(targets, user, itemIndex):
	global battleTurn
	global playersHaveMoved
	global enemiesHaveMoved

	clearScreen()
	renderBattleStatusMenu()

	# Get user text for strings
	userName = user.name
	targetNames = []
	for i in range(len(targets)):
		targetNames.append(targets[i].name)

	chosenItem = itemsList[itemIndex]

	# Reduce item amount by 1
	currentInventory[itemIndex] -= 1

	# String error checking
	indefiniteArticle = "an" if (chosenItem.name[0].lower() in ["a", "e", "i", "o", "u"]) else "a"
	scrollingPrint("{0} used {1} {2}.".format(userName, indefiniteArticle, chosenItem.name))

	currentEventText = ""
	# If item is of health type
	if (chosenItem.itemType == itemTypes[0]):
		allRecovery = []
		for i in range(len(targets)):
			currentRecovery = max(chosenItem.rawPower, round((chosenItem.relativePower / 100) * targets[i].maxHP))
			targets[i].curHP += currentRecovery
			allRecovery.append(currentRecovery)

		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			currentEventText += "{0} recovered {1} HP for {2}".format(currentUserName, allRecovery[i], targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."

	# Else, if item is of mana type
	elif (chosenItem.itemType == itemTypes[1]):
		allRecovery = []
		for i in range(len(targets)):
			currentRecovery = round((chosenItem.relativePower / 100) * targets[i].maxMP)
			targets[i].curMP += currentRecovery
			allRecovery.append(currentRecovery)

		for i in range(len(targets)):
			currentUserName = userName if (i == 0) else formatText("and ", 2, len(userName)) if (i + 1 == len(targets)) else " " * len(userName)

			currentEventText += "{0} recovered {1} MP for {2}".format(currentUserName, allRecovery[i], targetNames[i])

			if (i + 1 < len(targets)):
				currentEventText += ",\n"
			else:
				currentEventText += "."

	# # Else, item is of damage type
	# else:
	# 	damageMod = 1

	# 	if skill.statType == statTypes[1]:
	# 		userAttackStat = user.totalStats.ballisticAttack
	# 		targetDefenseStat = target.totalStats.ballisticDefense
	# 	elif skill.statType == statTypes[2]:
	# 		userAttackStat = user.totalStats.magicAttack
	# 		targetDefenseStat = target.totalStats.magicDefense
	scrollingPrint(currentEventText)

	user.evaluateCurrentPoints()
	for i in range(len(targets)):
		targets[i].evaluateCurrentPoints()

	if (user in currentPlayers):
		playersHaveMoved[selectedPlayer] = True
	elif (user in currentEnemies):
		enemiesHaveMoved[selectedEnemy] = True


renderMainMenu()
sys.exit()